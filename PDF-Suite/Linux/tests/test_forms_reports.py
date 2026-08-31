"""
PT-PT: Testes da deteccao de campos, da escrita do AcroForm, da leitura de
       documentos e dos relatorios.

       Ao contrario dos outros ficheiros de teste, este cria PDF a serio e
       volta a le-los. E a unica forma de verificar que um formulario gerado
       abre mesmo: um AcroForm mal formado nao levanta excepcao nenhuma ao ser
       escrito — so nao funciona quando alguem o abre, que ja e tarde.

EN-UK: Tests for field detection, AcroForm writing, document reading and
       reports.

       Unlike the other test files, this one creates real PDFs and reads them
       back. It is the only way to check that a generated form actually opens: a
       malformed AcroForm raises nothing when written — it simply fails when
       somebody opens it, which is too late.

Created by Redfox using Claude
"""

from __future__ import annotations

from pathlib import Path

import pytest

from pdfsuite import extract, reports
from pdfsuite.models import Campo, Documento, TipoCampo, nome_seguro_campo


@pytest.fixture
def formulario_pdf(tmp_path: Path) -> Path:
    """
    PT-PT: Um formulario em papel, com linha, sublinhado, quadrado e caixa.
    EN-UK: A paper form with a rule, an underscore run, a square and a box.
    """
    from reportlab.lib.pagesizes import A4
    from reportlab.pdfgen import canvas

    caminho = tmp_path / "formulario.pdf"
    c = canvas.Canvas(str(caminho), pagesize=A4)
    largura, altura = A4

    c.setFont("Helvetica", 11)

    c.drawString(60, altura - 100, "Nome completo:")
    c.line(160, altura - 103, 400, altura - 103)

    c.drawString(60, altura - 130, "Email: ______________________")

    c.rect(60, altura - 165, 10, 10)
    c.drawString(78, altura - 162, "Aceito os termos")

    c.drawString(60, altura - 200, "Observações:")
    c.rect(60, altura - 290, 400, 80)

    c.line(60, altura - 340, 260, altura - 340)
    c.setFont("Helvetica", 8)
    c.drawString(60, altura - 352, "Assinatura do requerente")

    c.showPage()
    c.save()
    return caminho


class TestNomeSeguroCampo:
    def test_normaliza_acentos_e_espacos(self):
        assert nome_seguro_campo("Data de início", set()) == "data_de_inicio"

    def test_remove_o_ponto(self):
        """
        PT-PT: O ponto é reservado: num PDF separa campo de campo-pai numa
               hierarquia, portanto «Data.Nascimento» criaria uma árvore em vez
               de um campo com esse nome.
        EN-UK: The full stop is reserved: in a PDF it separates a field from its
               parent in a hierarchy.
        """
        assert "." not in nome_seguro_campo("Data.Nascimento", set())

    def test_duplicados_sao_numerados(self):
        """
        PT-PT: Dois campos com o mesmo nome num AcroForm não são dois campos:
               são o mesmo campo em dois sítios, e escrever num escreve no
               outro. Num formulário com «Nome» em três páginas, é um bug que
               só aparece depois de alguém o preencher.
        EN-UK: Two fields sharing a name are one field in two places.
        """
        usados: set[str] = set()
        assert nome_seguro_campo("Nome", usados) == "nome"
        assert nome_seguro_campo("Nome", usados) == "nome_2"
        assert nome_seguro_campo("Nome", usados) == "nome_3"

    def test_etiqueta_vazia(self):
        assert nome_seguro_campo("", set()) == "campo"

    def test_etiqueta_so_com_simbolos(self):
        assert nome_seguro_campo("--- ///", set()) == "campo"


class TestCampo:
    def test_normalizar_troca_coordenadas_invertidas(self):
        """PT-PT: Rectângulo arrastado ao contrário. / EN-UK: Dragged backwards."""
        campo = Campo(nome="x", pagina=0, x0=200, y0=300, x1=100, y1=200)
        campo.normalizar()
        assert campo.x0 == 100
        assert campo.y0 == 200

    def test_campo_minusculo_e_invalido(self):
        assert not Campo(nome="x", pagina=0, x0=0, y0=0, x1=2, y1=2).valido()

    def test_campo_normal_e_valido(self):
        assert Campo(nome="x", pagina=0, x0=0, y0=0, x1=200, y1=16).valido()


class TestDeteccao:
    def test_encontra_os_campos_do_formulario(self, formulario_pdf: Path):
        from pdfsuite.detect import detectar

        campos, _ = detectar(formulario_pdf)
        nomes = {c.nome for c in campos}

        assert "nome_completo" in nomes
        assert "email" in nomes
        assert any("assinatura" in n for n in nomes)

    def test_quadrado_pequeno_e_caixa_de_seleccao(self, formulario_pdf: Path):
        from pdfsuite.detect import detectar

        campos, _ = detectar(formulario_pdf)
        caixas = [c for c in campos if c.tipo is TipoCampo.CAIXA]
        assert caixas
        assert any("aceito" in c.etiqueta.lower() for c in caixas)

    def test_caixa_grande_e_multilinha(self, formulario_pdf: Path):
        from pdfsuite.detect import detectar

        campos, _ = detectar(formulario_pdf)
        assert any(c.tipo is TipoCampo.MULTILINHA for c in campos)

    def test_etiqueta_por_baixo_da_linha(self, formulario_pdf: Path):
        """
        PT-PT: É a convenção das assinaturas: desenha-se a linha e escreve-se
               por baixo. Sem esta procura, todas as linhas de assinatura de
               todos os formulários ficavam sem nome.
        EN-UK: The signature convention: draw the rule, caption it below.
        """
        from pdfsuite.detect import detectar

        campos, _ = detectar(formulario_pdf)
        assinaturas = [c for c in campos if c.tipo is TipoCampo.ASSINATURA]
        assert assinaturas
        assert "equerente" in assinaturas[0].etiqueta

    def test_nomes_sao_unicos(self, formulario_pdf: Path):
        from pdfsuite.detect import detectar

        campos, _ = detectar(formulario_pdf)
        nomes = [c.nome for c in campos]
        assert len(nomes) == len(set(nomes))

    def test_campos_nao_se_sobrepoem(self, formulario_pdf: Path):
        """
        PT-PT: Um campo desenhado como rectângulo tem quatro linhas, e o
               detector de linhas vê nele um campo por cada lado.
        EN-UK: A field drawn as a rectangle has four sides.
        """
        from pdfsuite.detect import _sobrepoe, detectar

        campos, _ = detectar(formulario_pdf)
        for i, primeiro in enumerate(campos):
            for segundo in campos[i + 1 :]:
                assert not _sobrepoe(primeiro, segundo)

    def test_coordenadas_dentro_da_pagina(self, formulario_pdf: Path):
        """
        PT-PT: O pdfplumber conta de cima para baixo e o PDF de baixo para
               cima. Trocar as convenções produz campos correctos na horizontal
               e invertidos na vertical.
        EN-UK: pdfplumber counts from the top, the PDF from the bottom.
        """
        from pdfsuite.detect import detectar

        campos, _ = detectar(formulario_pdf)
        for campo in campos:
            assert 0 <= campo.y0 < campo.y1 <= 842
            assert 0 <= campo.x0 < campo.x1 <= 596

    def test_desligar_dois_pontos_reduz_os_campos(self, formulario_pdf: Path):
        from pdfsuite.detect import detectar

        com, _ = detectar(formulario_pdf, usar_dois_pontos=True)
        sem, _ = detectar(formulario_pdf, usar_dois_pontos=False)
        assert len(sem) <= len(com)


class TestCriarFormulario:
    def test_grava_e_volta_a_ler(self, formulario_pdf: Path, tmp_path: Path):
        from pdfsuite.detect import detectar
        from pdfsuite.forms import criar_formulario, listar_campos, tem_formulario

        campos, _ = detectar(formulario_pdf)
        destino = tmp_path / "preenchivel.pdf"
        escritos, _ = criar_formulario(formulario_pdf, destino, campos)

        assert escritos == len(campos)
        assert tem_formulario(destino) == escritos
        assert len(listar_campos(destino)) == escritos

    def test_original_nao_e_alterado(self, formulario_pdf: Path, tmp_path: Path):
        """
        PT-PT: A aplicação grava sempre uma cópia. O original de um formulário
               oficial não pode ser tocado.
        EN-UK: The application always writes a copy.
        """
        from pdfsuite.detect import detectar
        from pdfsuite.forms import criar_formulario, tem_formulario

        antes = formulario_pdf.read_bytes()
        campos, _ = detectar(formulario_pdf)
        criar_formulario(formulario_pdf, tmp_path / "novo.pdf", campos)

        assert formulario_pdf.read_bytes() == antes
        assert tem_formulario(formulario_pdf) == 0

    def test_preencher_e_reler(self, formulario_pdf: Path, tmp_path: Path):
        from pdfsuite.detect import detectar
        from pdfsuite.forms import criar_formulario, listar_campos, preencher

        campos, _ = detectar(formulario_pdf)
        vazio = tmp_path / "vazio.pdf"
        criar_formulario(formulario_pdf, vazio, campos)

        cheio = tmp_path / "cheio.pdf"
        preencher(vazio, cheio, {"nome_completo": "Rafael Santos"})

        valores = {n: v for n, _, v in listar_campos(cheio)}
        assert valores["nome_completo"] == "Rafael Santos"

    def test_caixa_de_seleccao_tem_os_dois_estados(self, formulario_pdf: Path, tmp_path: Path):
        """
        PT-PT: Uma caixa sem estado «Off» declarado não se consegue desmarcar
               depois de marcada.
        EN-UK: A tick box with no declared "Off" state cannot be unticked.
        """
        from pypdf import PdfReader

        from pdfsuite.detect import detectar
        from pdfsuite.forms import criar_formulario

        campos, _ = detectar(formulario_pdf)
        destino = tmp_path / "caixas.pdf"
        criar_formulario(formulario_pdf, destino, campos)

        leitor = PdfReader(str(destino))
        botoes = [
            anotacao.get_object()
            for pagina in leitor.pages
            for anotacao in pagina.get("/Annots", [])
            if anotacao.get_object().get("/FT") == "/Btn"
        ]
        assert botoes
        for botao in botoes:
            estados = botao["/AP"]["/N"]
            assert "/Off" in estados
            assert "/Sim" in estados

    def test_campos_sao_imprimiveis(self, formulario_pdf: Path, tmp_path: Path):
        """
        PT-PT: Sem a bandeira de impressão, o campo aparece no ecrã e
               desaparece na impressão — o que num formulário preenchido é a
               pior falha possível, porque só se descobre depois de estar
               assinado e entregue em papel.
        EN-UK: Without the print flag, the field shows on screen and vanishes
               when printed.
        """
        from pypdf import PdfReader

        from pdfsuite.detect import detectar
        from pdfsuite.forms import F_IMPRIMIVEL, criar_formulario

        campos, _ = detectar(formulario_pdf)
        destino = tmp_path / "imprimir.pdf"
        criar_formulario(formulario_pdf, destino, campos)

        leitor = PdfReader(str(destino))
        for pagina in leitor.pages:
            for anotacao in pagina.get("/Annots", []):
                objecto = anotacao.get_object()
                if objecto.get("/Subtype") == "/Widget":
                    assert int(objecto.get("/F", 0)) & F_IMPRIMIVEL

    def test_recursos_declaram_as_duas_letras(self, formulario_pdf: Path, tmp_path: Path):
        """
        PT-PT: Um leitor que encontre `/Helv` sem a encontrar nos recursos ou
               não mostra o texto ou escolhe uma letra ao acaso. A ZapfDingbats
               é procurada por reflexo pelos leitores que regeneram as caixas.
        EN-UK: A reader meeting `/Helv` without finding it in the resources
               either shows no text or picks a font at random.
        """
        from pypdf import PdfReader

        from pdfsuite.detect import detectar
        from pdfsuite.forms import criar_formulario

        campos, _ = detectar(formulario_pdf)
        destino = tmp_path / "letras.pdf"
        criar_formulario(formulario_pdf, destino, campos)

        letras = PdfReader(str(destino)).trailer["/Root"]["/AcroForm"]["/DR"]["/Font"]
        assert "/Helv" in letras
        assert "/ZaDb" in letras

    def test_campo_fora_da_pagina_e_recusado(self, formulario_pdf: Path, tmp_path: Path):
        campo_bom = Campo(nome="bom", pagina=0, x0=50, y0=50, x1=250, y1=66)
        campo_mau = Campo(nome="mau", pagina=9, x0=50, y0=50, x1=250, y1=66)

        from pdfsuite.forms import criar_formulario

        escritos, avisos = criar_formulario(
            formulario_pdf, tmp_path / "parcial.pdf", [campo_bom, campo_mau]
        )
        assert escritos == 1
        assert any("mau" in a for a in avisos)

    def test_sem_campos_validos_levanta(self, formulario_pdf: Path, tmp_path: Path):
        from pdfsuite.forms import criar_formulario

        minusculo = Campo(nome="x", pagina=0, x0=0, y0=0, x1=1, y1=1)
        with pytest.raises(ValueError):
            criar_formulario(formulario_pdf, tmp_path / "nada.pdf", [minusculo])


class TestLeitura:
    def test_ficheiro_inexistente(self, tmp_path: Path):
        documento = extract.ler(tmp_path / "nao-existe.pdf")
        assert not documento.ok
        assert "não existe" in documento.erro

    def test_ficheiro_vazio(self, tmp_path: Path):
        vazio = tmp_path / "vazio.txt"
        vazio.write_text("", encoding="utf-8")
        assert "vazio" in extract.ler(vazio).erro

    def test_formato_nao_suportado(self, tmp_path: Path):
        ficheiro = tmp_path / "imagem.png"
        ficheiro.write_bytes(b"\x89PNG\r\n")
        assert "não suportado" in extract.ler(ficheiro).erro

    def test_doc_antigo_explica_o_que_fazer(self, tmp_path: Path):
        """
        PT-PT: O .doc antigo é um formato binário diferente do .docx. Dizer o
               que fazer a seguir vale mais do que um erro genérico.
        EN-UK: The old .doc is a different binary format.
        """
        ficheiro = tmp_path / "antigo.doc"
        ficheiro.write_bytes(b"\xd0\xcf\x11\xe0")
        assert ".docx" in extract.ler(ficheiro).erro

    def test_texto_utf8(self, tmp_path: Path):
        ficheiro = tmp_path / "nota.txt"
        ficheiro.write_text("Preço da proposta: 1.500,00 €", encoding="utf-8")
        assert "€" in extract.ler(ficheiro).texto

    def test_texto_cp1252(self, tmp_path: Path):
        """
        PT-PT: Ficheiros exportados por aplicações de gestão antigas em Windows.
        EN-UK: Files exported by older Windows business applications.
        """
        ficheiro = tmp_path / "antigo.txt"
        ficheiro.write_bytes("Proposta comercial — condições".encode("cp1252"))
        assert "condi" in extract.ler(ficheiro).texto

    def test_pdf_e_lido(self, formulario_pdf: Path):
        documento = extract.ler(formulario_pdf)
        assert documento.ok
        assert documento.paginas == 1
        assert "Nome completo" in documento.texto

    def test_pdf_sem_texto_e_marcado_como_digitalizado(self, tmp_path: Path):
        """
        PT-PT: Um documento vazio e um PDF digitalizado são os dois «sem
               texto», mas o primeiro não tem nada lá dentro e o segundo tem
               tudo — em imagem. Dizer «o ficheiro está vazio» quando o problema
               é falta de OCR manda o utilizador procurar no sítio errado.
        EN-UK: An empty document and a scanned PDF are both "no text".
        """
        from reportlab.lib.pagesizes import A4
        from reportlab.pdfgen import canvas

        caminho = tmp_path / "digitalizado.pdf"
        c = canvas.Canvas(str(caminho), pagesize=A4)
        c.rect(100, 100, 200, 200)
        c.showPage()
        c.save()

        documento = extract.ler(caminho)
        assert documento.digitalizado
        assert "OCR" in documento.erro

    def test_ler_varios_mantem_os_que_falham(self, tmp_path: Path, formulario_pdf: Path):
        """
        PT-PT: Numa comparação de seis propostas, ficar com cinco sem perceber
               qual faltou é pior do que não ter nenhuma.
        EN-UK: Ending up with five and not knowing which dropped out is worse.
        """
        documentos = extract.ler_varios([formulario_pdf, tmp_path / "fantasma.pdf"])
        assert len(documentos) == 2
        assert documentos[1].erro


class TestRelatorios:
    def _comparacao(self):
        from pdfsuite.models import Proposta, Valor
        from pdfsuite.scoring import comparar

        propostas = []
        for nome, total in (("Alfa <Lda>", 9510.0), ("Beta", 11485.0)):
            proposta = Proposta(documento=Documento(caminho=Path(f"{nome}.pdf"), texto="x"))
            proposta.fornecedor = Valor(valor=nome, confianca=1.0)
            proposta.total = Valor(valor=total, confianca=1.0)
            proposta.iva_incluido = False
            proposta.garantia_meses = Valor(valor=36.0, confianca=1.0)
            propostas.append(proposta)
        return comparar(propostas)

    def test_html_escapa_o_conteudo_dos_documentos(self):
        """
        PT-PT: Uma proposta em PDF traz nomes de artigos com sinais de menor e
               maior mais vezes do que se imagina — «rede <1Gbps>», «prazo
               <=30 dias» — e inseri-los em bruto parte a página. Se em vez de
               um fragmento inofensivo aparecesse um `<script>`, o relatório
               passava a executá-lo ao ser aberto.
        EN-UK: A PDF quote carries item names with angle brackets more often
               than one imagines.
        """
        html = reports.relatorio_comparacao(self._comparacao())
        assert "<Lda>" not in html
        assert "&lt;Lda&gt;" in html

    def test_html_e_um_documento_completo(self):
        html = reports.relatorio_comparacao(self._comparacao())
        assert html.startswith("<!DOCTYPE html>")
        assert html.rstrip().endswith("</html>")
        assert 'charset="utf-8"' in html

    def test_excel_tem_as_folhas_esperadas(self, tmp_path: Path):
        from openpyxl import load_workbook

        destino = reports.excel_comparacao(self._comparacao(), tmp_path / "c.xlsx")
        livro = load_workbook(destino)
        assert "Comparação" in livro.sheetnames

    def test_excel_leva_formulas_vivas(self, tmp_path: Path):
        """
        PT-PT: Quem recebe a comparação quase sempre quer mexer nos pesos. Com
               os valores calculados e colados, teria de refazer as contas à
               mão; com fórmulas, muda o peso e a folha repontua.
        EN-UK: Whoever receives the comparison almost always wants to change the
               weights.
        """
        from openpyxl import load_workbook

        destino = reports.excel_comparacao(self._comparacao(), tmp_path / "c.xlsx")
        folha = load_workbook(destino)["Comparação"]
        formulas = [
            c.value
            for linha in folha.iter_rows()
            for c in linha
            if isinstance(c.value, str) and c.value.startswith("=")
        ]
        assert formulas

    def test_gravar_html_nunca_sobrepoe(self, tmp_path: Path):
        """
        PT-PT: Duas comparações seguidas não podem perder a primeira — que é
               muitas vezes a que interessa, tirada antes de mexer nos pesos.
        EN-UK: Two comparisons in a row must not lose the first.
        """
        primeiro = reports.gravar_html("<html></html>", tmp_path, "comparacao")
        segundo = reports.gravar_html("<html></html>", tmp_path, "comparacao")
        assert primeiro != segundo
        assert len(list(tmp_path.glob("*.html"))) == 2

    def test_nome_seguro_remove_caracteres_proibidos(self):
        for proibido in '<>:"/\\|?*':
            assert proibido not in reports.nome_seguro(f"a{proibido}b")

    def test_nome_seguro_nunca_devolve_vazio(self):
        assert reports.nome_seguro("///") != ""


class TestResumo:
    def test_resume_um_documento(self):
        from pdfsuite.summarise import resumir

        texto = (
            "O relatório analisa a renovação da rede do edifício principal. "
            "A infraestrutura actual tem oito anos e os switches deixaram de ter "
            "suporte do fabricante em Março de 2026. "
            "O investimento estimado é de 45.000 € distribuídos por duas fases. "
            "A primeira fase substitui o core e está prevista para 30 dias. "
            "A segunda fase cobre os pisos de quartos e depende da ocupação. "
            "Recomenda-se avançar com a primeira fase ainda este trimestre."
        )
        resumo = resumir(Documento(caminho=Path("r.pdf"), texto=texto), frases_desejadas=3)

        assert len(resumo.frases) == 3
        assert all(f in texto for f in resumo.frases)

    def test_frases_vem_por_ordem_de_leitura(self):
        """
        PT-PT: Um resumo cujas frases aparecem por ordem de relevância lê-se
               como uma lista solta; por ordem de leitura, lê-se como um texto.
        EN-UK: In relevance order a summary reads as a loose list.
        """
        from pdfsuite.summarise import dividir_frases, resumir

        texto = " ".join(
            f"Esta é a frase número {i} do relatório sobre a rede e os switches."
            for i in range(1, 12)
        )
        documento = Documento(caminho=Path("r.pdf"), texto=texto)
        resumo = resumir(documento, frases_desejadas=4)

        todas = dividir_frases(texto)
        posicoes = [todas.index(f) for f in resumo.frases]
        assert posicoes == sorted(posicoes)

    def test_documento_ilegivel_devolve_resumo_vazio(self):
        documento = Documento(caminho=Path("x.pdf"), erro="digitalizado")
        from pdfsuite.summarise import resumir

        assert resumir(documento).frases == []

    def test_termos_exclusivos(self):
        """
        PT-PT: Serve para o caso em que a comparação não é de preços: seis
               relatórios sobre o mesmo assunto, e a pergunta é o que um diz
               que os outros não dizem.
        EN-UK: For when the comparison is not about prices.
        """
        from pdfsuite.summarise import comparar_textos

        documentos = [
            Documento(caminho=Path("a.pdf"), texto="switches routers firewall cabolagem"),
            Documento(caminho=Path("b.pdf"), texto="switches routers firewall wireless"),
        ]
        resultado = comparar_textos(documentos)

        assert "switches" in resultado["__comuns__"]
        assert "cabolagem" in resultado["a"]
        assert "wireless" in resultado["b"]
