# -*- coding: utf-8 -*-
"""
PT-PT: Relatorios em HTML e em Excel.

       Sao dois formatos porque servem duas coisas diferentes. O HTML e para
       ler e anexar a um pedido de aprovacao: tem o veredicto, os avisos e o
       raciocinio. O Excel e para trabalhar: quem recebe o relatorio quase
       sempre quer mexer nos pesos, acrescentar uma coluna ou juntar isto a
       outra folha.

       Como em qualquer relatorio gerado a partir de ficheiros de terceiros,
       tudo o que vem dos documentos passa por `escape()` antes de entrar no
       HTML. Uma proposta comercial em PDF traz nomes de artigos com sinais de
       menor e maior mais vezes do que se imagina — «rede <1Gbps>», «prazo
       <=30 dias» — e insere-los em bruto parte a pagina.

EN-UK: HTML and Excel reports.

       Two formats because they serve two different things. The HTML is to read
       and attach to an approval request; the Excel is to work with, because
       whoever receives the report almost always wants to change the weights or
       add a column.

       As in any report generated from third-party files, everything coming
       from the documents goes through `escape()` before entering the HTML.

Created by Redfox using Claude
"""

from __future__ import annotations

import datetime as dt
import logging
import re
from html import escape
from pathlib import Path

from . import __credit__, __version__
from .models import Campo, Comparacao, Resumo
from .money import formatar_moeda
from .scoring import poupanca

log = logging.getLogger(__name__)

_CSS = """
:root { color-scheme: light; }
* { box-sizing: border-box; }
body {
  font-family: "Segoe UI", system-ui, -apple-system, sans-serif;
  margin: 0; padding: 32px 24px; background: #f4f5f7; color: #1c2833;
  line-height: 1.55;
}
.folha { max-width: 1120px; margin: 0 auto; }
header { border-bottom: 3px solid #1c2833; padding-bottom: 16px; margin-bottom: 24px; }
h1 { font-size: 24px; margin: 0 0 4px; }
h2 { font-size: 18px; margin: 32px 0 12px; }
h3 { font-size: 15px; margin: 0 0 6px; }
.sub { color: #5d6d7e; font-size: 13px; }
.veredicto {
  background: #fff; border-left: 5px solid #2d6a4f; border-radius: 4px;
  padding: 16px 20px; margin: 20px 0; font-size: 16px;
}
.veredicto.incerto { border-left-color: #a8620c; }
.veredicto b { font-size: 19px; }
.aviso {
  background: #fdf3e3; border-left: 5px solid #a8620c; border-radius: 4px;
  padding: 11px 16px; margin: 8px 0; font-size: 14px;
}
.nota {
  background: #eef2f7; border-left: 4px solid #5d6d7e; border-radius: 4px;
  padding: 10px 14px; margin: 6px 0; font-size: 13px; color: #34495e;
}
.cartoes { display: flex; flex-wrap: wrap; gap: 12px; margin: 20px 0; }
.cartao {
  background: #fff; border: 1px solid #d5d8dc; border-radius: 6px;
  padding: 14px 18px; min-width: 150px;
}
.cartao .n { font-size: 22px; font-weight: 600; }
.cartao .r { font-size: 11px; color: #5d6d7e; text-transform: uppercase; letter-spacing: .4px; }
table {
  border-collapse: collapse; width: 100%; background: #fff;
  border: 1px solid #d5d8dc; border-radius: 4px; overflow: hidden;
  font-size: 14px; margin-bottom: 8px;
}
th, td { text-align: left; padding: 8px 11px; border-bottom: 1px solid #eaecee; }
th {
  background: #eaecee; font-size: 11px; text-transform: uppercase;
  letter-spacing: .4px; color: #5d6d7e; white-space: nowrap;
}
td.num, th.num { text-align: right; font-variant-numeric: tabular-nums; }
tr:last-child td { border-bottom: none; }
tr.vencedora td { background: #eaf5ef; font-weight: 600; }
.barra {
  display: inline-block; height: 8px; border-radius: 4px;
  background: #33477a; vertical-align: middle; margin-right: 6px;
}
.faltou { color: #a8620c; font-style: italic; }
.item {
  background: #fff; border: 1px solid #d5d8dc; border-left: 5px solid #33477a;
  border-radius: 4px; padding: 14px 18px; margin-bottom: 10px;
}
blockquote {
  margin: 8px 0; padding: 8px 14px; border-left: 3px solid #d5d8dc;
  color: #34495e; font-size: 14px;
}
pre {
  background: #f4f5f7; border: 1px solid #e5e7e9; border-radius: 3px;
  padding: 10px 12px; font-size: 12px; white-space: pre-wrap;
  word-break: break-word; margin: 6px 0;
  font-family: Consolas, "Courier New", monospace;
}
.etiquetas span {
  display: inline-block; background: #eaecee; border-radius: 3px;
  padding: 2px 8px; margin: 2px 4px 2px 0; font-size: 12px;
}
footer {
  margin-top: 40px; padding-top: 16px; border-top: 1px solid #d5d8dc;
  font-size: 12px; color: #5d6d7e;
}
@media print { body { background: #fff; padding: 0; } table, .item { break-inside: avoid; } }
"""


def nome_seguro(bruto: str) -> str:
    """
    PT-PT: Reduz um texto a um nome de ficheiro valido em Windows.
    EN-UK: Reduces text to a valid Windows file name.
    """
    limpo = re.sub(r'[<>:"/\\|?*\x00-\x1f]', "", bruto or "").strip(" .")
    limpo = re.sub(r"\s+", "_", limpo)
    return limpo[:60] or "relatorio"


def _documento(titulo: str, corpo: str) -> str:
    """PT-PT: Envolve o corpo em HTML completo. / EN-UK: Wraps in full HTML."""
    return (
        "<!DOCTYPE html>\n"
        '<html lang="pt-PT"><head><meta charset="utf-8">'
        '<meta name="viewport" content="width=device-width, initial-scale=1">'
        f"<title>{escape(titulo)}</title><style>{_CSS}</style></head>"
        f'<body><div class="folha">{corpo}'
        f"<footer>{escape(__credit__)} · PDF Suite {escape(__version__)}</footer>"
        "</div></body></html>"
    )


def _cabecalho(titulo: str, subtitulo: str) -> str:
    agora = dt.datetime.now().strftime("%d/%m/%Y às %H:%M")
    return (
        "<header>"
        f"<h1>{escape(titulo)}</h1>"
        f'<div class="sub">{escape(subtitulo)}</div>'
        f'<div class="sub">Gerado em {escape(agora)}</div>'
        "</header>"
    )


# ---------------------------------------------------------------------------
# PT-PT: Relatorio de comparacao / EN-UK: Comparison report
# ---------------------------------------------------------------------------


def relatorio_comparacao(comparacao: Comparacao) -> str:
    """
    PT-PT: Relatorio HTML da comparacao de propostas.
    EN-UK: HTML report of the proposal comparison.
    """
    corpo = [
        _cabecalho(
            "Comparação de Propostas",
            f"{len(comparacao.pontuacoes)} proposta(s) analisada(s)",
        )
    ]

    ordem = comparacao.ordenadas
    if not ordem:
        corpo.append(
            '<div class="aviso">Não foi possível analisar nenhuma proposta.</div>'
        )
        for aviso in comparacao.avisos:
            corpo.append(f'<div class="aviso">{escape(aviso)}</div>')
        return _documento("Comparação de Propostas", "".join(corpo))

    primeira = ordem[0]

    # --- PT-PT: Veredicto / EN-UK: Verdict --------------------------------
    if comparacao.decisao_segura:
        texto = (
            f"Melhor pontuada: <b>{escape(primeira.proposta.rotulo)}</b>, "
            f"com {primeira.total:.1f} pontos em 100."
        )
        if len(ordem) > 1:
            texto += (
                f" A seguinte, {escape(ordem[1].proposta.rotulo)}, fica "
                f"{primeira.total - ordem[1].total:.1f} pontos atrás."
            )
        classe = "veredicto"
    else:
        texto = (
            f"<b>Não há vencedor claro.</b> "
            f"{escape(primeira.proposta.rotulo)} e "
            f"{escape(ordem[1].proposta.rotulo) if len(ordem) > 1 else '—'} "
            "estão dentro da margem de erro da extracção automática. "
            "A decisão tem de ser tomada com critérios que esta análise não mede."
        )
        classe = "veredicto incerto"

    corpo.append(f'<div class="{classe}">{texto}</div>')

    # --- PT-PT: Cartoes de resumo / EN-UK: Summary cards -------------------
    diferenca = poupanca(comparacao)
    corpo.append('<div class="cartoes">')
    corpo.append(
        f'<div class="cartao"><div class="n">{len(ordem)}</div>'
        '<div class="r">propostas</div></div>'
    )
    if diferenca:
        valor, barata, cara = diferenca
        corpo.append(
            f'<div class="cartao"><div class="n">{escape(formatar_moeda(valor))}</div>'
            '<div class="r">entre a mais barata e a mais cara</div></div>'
        )
        corpo.append(
            f'<div class="cartao"><div class="n" style="font-size:16px">{escape(barata)}</div>'
            '<div class="r">preço mais baixo</div></div>'
        )
    corpo.append(
        f'<div class="cartao"><div class="n">{comparacao.taxa_iva_omissao:g}%</div>'
        '<div class="r">IVA assumido quando não declarado</div></div>'
    )
    corpo.append("</div>")

    for aviso in comparacao.avisos:
        corpo.append(f'<div class="aviso">{escape(aviso)}</div>')

    # --- PT-PT: Tabela de pontuacao / EN-UK: Score table -------------------
    corpo.append("<h2>Pontuação</h2>")
    corpo.append("<table><tr><th>#</th><th>Proposta</th><th class='num'>Pontos</th>")
    for criterio in comparacao.criterios:
        corpo.append(
            f"<th class='num'>{escape(criterio.etiqueta)}"
            f"<br><span style='font-weight:400;text-transform:none'>"
            f"peso {criterio.peso:g}</span></th>"
        )
    corpo.append("<th class='num'>Dados</th></tr>")

    for posicao, pontuacao in enumerate(ordem, 1):
        classe = " class='vencedora'" if posicao == 1 and comparacao.decisao_segura else ""
        largura = max(pontuacao.total, 1) * 0.9
        corpo.append(
            f"<tr{classe}><td>{posicao}</td>"
            f"<td>{escape(pontuacao.proposta.rotulo)}</td>"
            f"<td class='num'><span class='barra' style='width:{largura:.0f}px'></span>"
            f"{pontuacao.total:.1f}</td>"
        )
        for criterio in comparacao.criterios:
            valor = pontuacao.valores.get(criterio.chave)
            if valor is None:
                corpo.append("<td class='num faltou'>não diz</td>")
            else:
                if criterio.chave == "preco":
                    apresentado = formatar_moeda(valor, pontuacao.proposta.moeda or "EUR")
                else:
                    apresentado = f"{valor:g} {criterio.unidade}".strip()
                pontos = pontuacao.por_criterio.get(criterio.chave, 0)
                corpo.append(
                    f"<td class='num'>{escape(apresentado)}"
                    f"<br><span class='sub'>{pontos:.0f} pts</span></td>"
                )
        corpo.append(f"<td class='num'>{pontuacao.completude:.0f}%</td></tr>")
    corpo.append("</table>")
    corpo.append(
        '<div class="sub">A coluna «Dados» diz quantos critérios a proposta '
        "declara. Uma pontuação alta com poucos dados assenta em poucos critérios "
        "e não é comparável com uma pontuação alta com dados completos.</div>"
    )

    # --- PT-PT: Detalhe por proposta / EN-UK: Per-proposal detail ----------
    corpo.append("<h2>Detalhe por proposta</h2>")
    for posicao, pontuacao in enumerate(ordem, 1):
        proposta = pontuacao.proposta
        corpo.append('<div class="item">')
        corpo.append(
            f"<h3>{posicao}. {escape(proposta.rotulo)} — {pontuacao.total:.1f} pontos</h3>"
        )
        corpo.append(
            f'<div class="sub">Ficheiro: {escape(proposta.documento.nome)}'
            + (
                f" · Referência: {escape(str(proposta.referencia.valor))}"
                if proposta.referencia.conhecido
                else ""
            )
            + "</div>"
        )

        if proposta.total.conhecido:
            iva = {
                True: "IVA incluído",
                False: "acresce IVA",
                None: "IVA não declarado",
            }[proposta.iva_incluido]
            corpo.append(
                "<table><tr><th>Total no documento</th><th>Regime de IVA</th>"
                "<th>Total comparável</th></tr>"
                f"<tr><td class='num'>{escape(formatar_moeda(proposta.total.valor, proposta.moeda or 'EUR'))}</td>"
                f"<td>{escape(iva)}</td>"
                f"<td class='num'><b>{escape(formatar_moeda(proposta.total_com_iva(comparacao.taxa_iva_omissao), proposta.moeda or 'EUR'))}</b></td>"
                "</tr></table>"
            )
            if proposta.total.contexto:
                corpo.append(
                    "<blockquote>Lido de: "
                    f"…{escape(proposta.total.contexto)}…</blockquote>"
                )

        for nota in proposta.notas:
            corpo.append(f'<div class="nota">{escape(nota)}</div>')

        corpo.append("</div>")

    corpo.append(
        '<div class="nota">Os valores foram extraídos automaticamente do texto '
        "dos documentos. A extracção acerta na maioria dos casos e falha em "
        "alguns, e as falhas não são visíveis no resultado final — confirme os "
        "totais no PDF original antes de adjudicar.</div>"
    )

    return _documento("Comparação de Propostas", "".join(corpo))


# ---------------------------------------------------------------------------
# PT-PT: Relatorio de resumo / EN-UK: Summary report
# ---------------------------------------------------------------------------


def relatorio_resumo(resumos: list[Resumo], comparacao_termos: dict | None = None) -> str:
    """
    PT-PT: Relatorio HTML de um ou varios resumos.
    EN-UK: HTML report of one or several summaries.
    """
    plural = "documentos" if len(resumos) != 1 else "documento"
    corpo = [_cabecalho("Resumo de Documentos", f"{len(resumos)} {plural}")]

    for resumo in resumos:
        corpo.append('<div class="item">')
        corpo.append(f"<h3>{escape(resumo.documento.nome)}</h3>")
        detalhes = [resumo.documento.formato, f"{resumo.documento.palavras} palavras"]
        if resumo.documento.paginas:
            detalhes.insert(1, f"{resumo.documento.paginas} páginas")
        corpo.append(f'<div class="sub">{escape(" · ".join(detalhes))}</div>')

        if resumo.documento.erro:
            corpo.append(f'<div class="aviso">{escape(resumo.documento.erro)}</div>')
            corpo.append("</div>")
            continue

        if resumo.documento.digitalizado:
            corpo.append(
                '<div class="aviso">O documento parece digitalizado: foi lido '
                "pouco texto e o resumo pode estar incompleto.</div>"
            )

        if resumo.frases:
            corpo.append("<p>" + escape(" ".join(resumo.frases)) + "</p>")

        if resumo.numeros:
            corpo.append(
                '<div class="sub" style="margin-top:10px">Valores e prazos no documento</div>'
                '<div class="etiquetas">'
                + "".join(f"<span>{escape(n)}</span>" for n in resumo.numeros)
                + "</div>"
            )

        if resumo.datas:
            corpo.append(
                '<div class="sub" style="margin-top:8px">Datas</div>'
                '<div class="etiquetas">'
                + "".join(f"<span>{escape(d)}</span>" for d in resumo.datas)
                + "</div>"
            )

        if resumo.palavras_chave:
            corpo.append(
                '<div class="sub" style="margin-top:8px">Termos mais frequentes</div>'
                '<div class="etiquetas">'
                + "".join(
                    f"<span>{escape(p)} <b>{n}</b></span>"
                    for p, n in resumo.palavras_chave[:12]
                )
                + "</div>"
            )

        if resumo.texto_ia:
            corpo.append(
                '<div class="nota"><b>Análise assistida por modelo</b> — o texto '
                "seguinte foi gerado por um modelo de linguagem a partir do "
                "documento, e não é uma citação dele.</div>"
                f"<pre>{escape(resumo.texto_ia)}</pre>"
            )

        corpo.append("</div>")

    if comparacao_termos:
        corpo.append("<h2>Termos exclusivos de cada documento</h2>")
        corpo.append(
            '<div class="sub">O que cada documento diz e os outros não. '
            "É o caminho mais curto para ver onde divergem.</div>"
        )
        comuns = comparacao_termos.get("__comuns__", [])
        if comuns:
            corpo.append(
                '<div class="item"><h3>Comuns a todos</h3><div class="etiquetas">'
                + "".join(f"<span>{escape(t)}</span>" for t in comuns)
                + "</div></div>"
            )
        for rotulo, termos in comparacao_termos.items():
            if rotulo == "__comuns__":
                continue
            corpo.append(
                f'<div class="item"><h3>{escape(rotulo)}</h3><div class="etiquetas">'
                + (
                    "".join(f"<span>{escape(t)}</span>" for t in termos)
                    or '<span class="faltou">nada exclusivo</span>'
                )
                + "</div></div>"
            )

    corpo.append(
        '<div class="nota">O resumo é extractivo: as frases apresentadas estão '
        "no documento tal e qual. Nenhum texto foi gerado, excepto onde estiver "
        "assinalado como análise assistida.</div>"
    )

    return _documento("Resumo de Documentos", "".join(corpo))


# ---------------------------------------------------------------------------
# PT-PT: Relatorio de formulario / EN-UK: Form report
# ---------------------------------------------------------------------------


def relatorio_formulario(origem: Path, destino: Path, campos: list[Campo], avisos: list[str]) -> str:
    """
    PT-PT: Ficha do formulario gerado, com a lista de campos.

           Serve de documentacao para quem depois quiser preencher o formulario
           por script: os nomes dos campos sao as chaves, e sem esta lista a
           unica forma de os conhecer e abrir o PDF numa ferramenta que os saiba
           ler.

    EN-UK: A record of the generated form, listing its fields. It doubles as
           documentation for whoever later fills the form by script: the field
           names are the keys.
    """
    corpo = [
        _cabecalho(
            "Formulário Preenchível",
            f"{len(campos)} campo(s) · gerado a partir de {origem.name}",
        )
    ]

    corpo.append(
        '<div class="veredicto">Ficheiro gerado: <b>'
        f"{escape(destino.name)}</b></div>"
    )

    for aviso in avisos:
        corpo.append(f'<div class="aviso">{escape(aviso)}</div>')

    por_pagina: dict[int, list[Campo]] = {}
    for campo in campos:
        por_pagina.setdefault(campo.pagina, []).append(campo)

    for pagina in sorted(por_pagina):
        corpo.append(f"<h2>Página {pagina + 1}</h2>")
        corpo.append(
            "<table><tr><th>Nome do campo</th><th>Etiqueta</th><th>Tipo</th>"
            "<th>Origem</th><th class='num'>Confiança</th></tr>"
        )
        for campo in por_pagina[pagina]:
            corpo.append(
                f"<tr><td><code>{escape(campo.nome)}</code></td>"
                f"<td>{escape(campo.etiqueta or '—')}</td>"
                f"<td>{escape(campo.tipo.etiqueta)}</td>"
                f"<td>{escape(campo.origem.value)}</td>"
                f"<td class='num'>{campo.confianca * 100:.0f}%</td></tr>"
            )
        corpo.append("</table>")

    corpo.append(
        '<div class="nota">Os nomes na primeira coluna são as chaves do '
        "formulário. Servem para preencher o PDF por script ou para importar os "
        "dados depois de preenchido.</div>"
    )

    return _documento("Formulário Preenchível", "".join(corpo))


# ---------------------------------------------------------------------------
# PT-PT: Excel / EN-UK: Excel
# ---------------------------------------------------------------------------


def _largura_automatica(folha) -> None:
    """
    PT-PT: Ajusta a largura das colunas ao conteudo.

           Sem isto, a coluna do fornecedor sai com oito caracteres e todos os
           nomes aparecem cortados — que e a primeira coisa que se nota ao abrir
           uma folha gerada por codigo, e a que mais depressa a faz parecer
           malfeita.

    EN-UK: Fits column widths to the content. Without it the vendor column comes
           out eight characters wide and every name is cut off, which is the
           first thing anyone notices on opening a generated sheet.
    """
    from openpyxl.utils import get_column_letter

    for indice, coluna in enumerate(folha.columns, 1):
        maior = max(
            (len(str(celula.value)) for celula in coluna if celula.value is not None),
            default=10,
        )
        folha.column_dimensions[get_column_letter(indice)].width = min(max(maior + 3, 11), 52)


def excel_comparacao(comparacao: Comparacao, destino: Path) -> Path:
    """
    PT-PT: Exporta a comparacao para Excel.

           Tres folhas: a comparacao com as formulas de pontuacao vivas, os
           dados em bruto e as notas.

           As formulas ficam vivas de proposito. E a diferenca entre uma folha
           que se le e uma folha com que se trabalha: mudar um peso na folha
           dos criterios recalcula a pontuacao toda, e quem tem de justificar a
           escolha numa reuniao precisa exactamente disso — mostrar que a
           conclusao aguenta pesos diferentes, ou onde e que deixa de aguentar.

    EN-UK: Exports the comparison to Excel. Three sheets: the comparison with
           live scoring formulas, the raw data, and the notes.

           The formulas are live on purpose. It is the difference between a
           sheet you read and a sheet you work with: changing a weight
           recalculates every score, and whoever has to justify the choice in a
           meeting needs exactly that.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    livro = Workbook()

    titulo_letra = Font(bold=True, size=11, color="FFFFFF")
    titulo_fundo = PatternFill("solid", fgColor="33477A")
    destaque = PatternFill("solid", fgColor="EAF5EF")
    borda = Border(bottom=Side(style="thin", color="D5D8DC"))
    centrado = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ordem = comparacao.ordenadas
    criterios = comparacao.criterios

    # --- PT-PT: Folha 1 — Comparação ---------------------------------------
    folha = livro.active
    folha.title = "Comparação"

    folha["A1"] = "Comparação de Propostas"
    folha["A1"].font = Font(bold=True, size=15)
    folha["A2"] = f"Gerado em {dt.datetime.now():%d/%m/%Y às %H:%M} · {__credit__}"
    folha["A2"].font = Font(size=9, color="5D6D7E")

    linha_criterios = 4
    folha.cell(row=linha_criterios, column=1, value="Pesos (editáveis)").font = Font(bold=True)
    for indice, criterio in enumerate(criterios):
        coluna = 2 + indice
        folha.cell(row=linha_criterios, column=coluna, value=criterio.etiqueta).font = Font(size=9)
        folha.cell(row=linha_criterios + 1, column=coluna, value=criterio.peso)
    folha.cell(row=linha_criterios + 1, column=1, value="→").font = Font(bold=True)

    cabecalho = linha_criterios + 3
    colunas = ["#", "Proposta", "Pontuação"]
    for criterio in criterios:
        colunas += [f"{criterio.etiqueta} (valor)", f"{criterio.etiqueta} (pts)"]
    colunas += ["Dados %", "Ficheiro"]

    for indice, nome in enumerate(colunas, 1):
        celula = folha.cell(row=cabecalho, column=indice, value=nome)
        celula.font = titulo_letra
        celula.fill = titulo_fundo
        celula.alignment = centrado

    for posicao, pontuacao in enumerate(ordem, 1):
        linha = cabecalho + posicao
        folha.cell(row=linha, column=1, value=posicao)
        folha.cell(row=linha, column=2, value=pontuacao.proposta.rotulo)

        colunas_pontos: list[str] = []
        colunas_pesos: list[str] = []

        for indice, criterio in enumerate(criterios):
            coluna_valor = 4 + indice * 2
            coluna_pontos = coluna_valor + 1

            valor = pontuacao.valores.get(criterio.chave)
            celula = folha.cell(row=linha, column=coluna_valor, value=valor)
            if valor is None:
                celula.value = "não diz"
                celula.font = Font(italic=True, color="A8620C")
            elif criterio.chave == "preco":
                celula.number_format = '#,##0.00 "€"'

            pontos = pontuacao.por_criterio.get(criterio.chave)
            folha.cell(row=linha, column=coluna_pontos, value=pontos)
            if pontos is not None:
                folha.cell(row=linha, column=coluna_pontos).number_format = "0.0"
                letra_pontos = get_column_letter(coluna_pontos)
                letra_peso = get_column_letter(2 + indice)
                colunas_pontos.append(f"{letra_pontos}{linha}*{letra_peso}${linha_criterios + 1}")
                colunas_pesos.append(f"{letra_peso}${linha_criterios + 1}")

        # PT-PT: A pontuacao e uma formula, nao um numero gravado. Mexer nos
        #        pesos em cima recalcula esta coluna.
        # EN-UK: The score is a formula, not a stored number.
        celula_total = folha.cell(row=linha, column=3)
        if colunas_pontos:
            celula_total.value = f"=({'+'.join(colunas_pontos)})/({'+'.join(colunas_pesos)})"
        else:
            celula_total.value = 0
        celula_total.number_format = "0.0"
        celula_total.font = Font(bold=True)

        folha.cell(
            row=linha, column=len(colunas) - 1, value=pontuacao.completude / 100
        ).number_format = "0%"
        folha.cell(row=linha, column=len(colunas), value=pontuacao.proposta.documento.nome)

        if posicao == 1 and comparacao.decisao_segura:
            for indice in range(1, len(colunas) + 1):
                folha.cell(row=linha, column=indice).fill = destaque

        for indice in range(1, len(colunas) + 1):
            folha.cell(row=linha, column=indice).border = borda

    folha.freeze_panes = folha.cell(row=cabecalho + 1, column=3)
    _largura_automatica(folha)

    # --- PT-PT: Folha 2 — Dados em bruto -----------------------------------
    dados = livro.create_sheet("Dados")
    colunas_dados = [
        "Proposta", "Ficheiro", "Referência", "Total no documento", "Moeda",
        "Regime de IVA", "Taxa de IVA", "Total comparável",
        "Pagamento (dias)", "Entrega (dias)", "Garantia (meses)", "Validade (dias)",
        "Confiança do total",
    ]
    for indice, nome in enumerate(colunas_dados, 1):
        celula = dados.cell(row=1, column=indice, value=nome)
        celula.font = titulo_letra
        celula.fill = titulo_fundo
        celula.alignment = centrado

    for posicao, pontuacao in enumerate(ordem, 2):
        proposta = pontuacao.proposta
        regime = {True: "incluído", False: "acresce", None: "não declarado"}[
            proposta.iva_incluido
        ]
        valores = [
            proposta.rotulo,
            proposta.documento.nome,
            proposta.referencia.valor if proposta.referencia.conhecido else "",
            proposta.total.valor if proposta.total.conhecido else "",
            proposta.moeda,
            regime,
            proposta.taxa_iva.valor if proposta.taxa_iva.conhecido else comparacao.taxa_iva_omissao,
            proposta.total_com_iva(comparacao.taxa_iva_omissao),
            proposta.prazo_pagamento.valor if proposta.prazo_pagamento.conhecido else "",
            proposta.prazo_entrega.valor if proposta.prazo_entrega.conhecido else "",
            proposta.garantia_meses.valor if proposta.garantia_meses.conhecido else "",
            proposta.validade.valor if proposta.validade.conhecido else "",
            proposta.total.confianca,
        ]
        for indice, valor in enumerate(valores, 1):
            celula = dados.cell(row=posicao, column=indice, value=valor)
            if indice in (4, 8):
                celula.number_format = '#,##0.00 "€"'
            if indice == 13:
                celula.number_format = "0%"

    dados.freeze_panes = "A2"
    _largura_automatica(dados)

    # --- PT-PT: Folha 3 — Notas e avisos -----------------------------------
    notas = livro.create_sheet("Notas")
    notas["A1"] = "Avisos da análise"
    notas["A1"].font = Font(bold=True, size=12)

    linha = 2
    for aviso in comparacao.avisos:
        notas.cell(row=linha, column=1, value=aviso).alignment = Alignment(wrap_text=True)
        linha += 1

    linha += 1
    notas.cell(row=linha, column=1, value="Notas por proposta").font = Font(bold=True, size=12)
    linha += 1

    for pontuacao in ordem:
        proposta = pontuacao.proposta
        if not proposta.notas:
            continue
        notas.cell(row=linha, column=1, value=proposta.rotulo).font = Font(bold=True)
        linha += 1
        for nota in proposta.notas:
            notas.cell(row=linha, column=2, value=nota).alignment = Alignment(wrap_text=True)
            linha += 1
        linha += 1

    linha += 1
    notas.cell(
        row=linha,
        column=1,
        value=(
            "Os valores foram extraídos automaticamente do texto dos documentos. "
            "A extracção acerta na maioria dos casos e falha em alguns, e as falhas "
            "não são visíveis no resultado final — confirme os totais no PDF "
            "original antes de adjudicar."
        ),
    ).alignment = Alignment(wrap_text=True)

    notas.column_dimensions["A"].width = 34
    notas.column_dimensions["B"].width = 92

    destino = Path(destino)
    destino.parent.mkdir(parents=True, exist_ok=True)
    livro.save(str(destino))
    log.info("Excel gravado em %s", destino)
    return destino


# ---------------------------------------------------------------------------
# PT-PT: Gravacao / EN-UK: Saving
# ---------------------------------------------------------------------------


def gravar_html(html: str, pasta: Path, prefixo: str) -> Path:
    """
    PT-PT: Grava o HTML com data e hora no nome.

           O contador existe porque o carimbo tem resolucao de um segundo e
           gerar dois relatorios seguidos leva menos do que isso — sem ele, o
           segundo apagava o primeiro.

    EN-UK: Writes the HTML with date and time in the name. The counter exists
           because the stamp has one-second resolution and producing two reports
           in a row takes less than that.
    """
    pasta = Path(pasta)
    pasta.mkdir(parents=True, exist_ok=True)

    carimbo = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    base = nome_seguro(prefixo)
    destino = pasta / f"{base}_{carimbo}.html"

    contador = 2
    while destino.exists():
        destino = pasta / f"{base}_{carimbo}_{contador}.html"
        contador += 1

    destino.write_text(html, encoding="utf-8")
    log.info("Relatório gravado em %s", destino)
    return destino


def caminho_livre(pasta: Path, nome: str, extensao: str) -> Path:
    """
    PT-PT: Um caminho que ainda nao existe, acrescentando um contador.

           Usado para os PDF gerados. Sobrepor em silencio o formulario que o
           utilizador acabou de rever e a forma mais rapida de lhe fazer perder
           o trabalho.

    EN-UK: A path that does not yet exist, adding a counter. Silently
           overwriting the form the user has just reviewed is the fastest way to
           lose their work.
    """
    pasta = Path(pasta)
    pasta.mkdir(parents=True, exist_ok=True)

    destino = pasta / f"{nome_seguro(nome)}{extensao}"
    contador = 2
    while destino.exists():
        destino = pasta / f"{nome_seguro(nome)}_{contador}{extensao}"
        contador += 1
    return destino
