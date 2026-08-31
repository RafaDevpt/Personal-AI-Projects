#!/usr/bin/env python3
"""
PT-PT: Exportação do mapa para Excel.

       O Excel é o destino real desta informação. Um PDF lê-se uma vez; uma
       folha filtra-se, ordena-se, e responde à pergunta que se tem naquele
       momento — "que impressoras há no piso 2", "que portas do SW-COZINHA
       estão livres", "onde está este MAC".

       Cinco folhas, cada uma com uma pergunta em mente:

       - **Resumo** — o que correu, quando, e o que ficou por fazer.
       - **Equipamentos** — os switches, e o que se conseguiu ler de cada um.
       - **Ligações** — o cabo a cabo entre switches.
       - **Pontos finais** — a folha grande, a que se filtra.
       - **Problemas** — o que vale a pena olhar.

       A coluna dos sinais na folha dos pontos finais é o que separa esta
       ferramenta de uma que adivinha: para cada classificação, o texto do que a
       sustentou. Quem discordar não tem de acreditar nem de ir à rede
       confirmar — lê a razão e decide.

EN-UK: Exporting the map to Excel.

       Excel is this information's real destination. A PDF is read once; a
       spreadsheet is filtered, sorted, and answers whatever question you have
       at that moment — "which printers are on floor 2", "which SW-COZINHA
       ports are free", "where is this MAC".

       Five sheets, each with a question in mind:

       - **Resumo** — what ran, when, and what was left undone.
       - **Equipamentos** — the switches, and what could be read from each.
       - **Ligações** — the cable-by-cable between switches.
       - **Pontos finais** — the big sheet, the one you filter.
       - **Problemas** — what is worth a look.

       The signals column on the endpoints sheet is what separates this tool
       from one that guesses: for every classification, the text of what
       supported it. Anyone who disagrees need not take it on faith nor go to
       the network to check — they read the reason and decide.

Created by Redfox using Claude
"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Any

from .. import __app_name__, __version__
from ..models import Confidence, Topology

# PT-PT: Cores por nível de confiança. Discretas de propósito: a folha é para
#        ler, não para piscar.
# EN-UK: Colours by confidence level. Deliberately muted: the sheet is for
#        reading, not for flashing.
_CONFIDENCE_FILL = {
    Confidence.HIGH: "E8F5EC",
    Confidence.MEDIUM: "FFF6E5",
    Confidence.LOW: "FDF0EE",
    Confidence.NONE: "F2F3F5",
}

_HEADER_FILL = "1F3A6E"


class ExcelError(RuntimeError):
    """PT-PT: Falha a escrever o Excel. / EN-UK: Failure writing the Excel file."""


def write(topology: Topology, path: Path, started: datetime | None = None) -> Path:
    """
    PT-PT: Escreve o mapa completo num livro de Excel.

    EN-UK: Writes the complete map into an Excel workbook.

    :param topology:
        PT-PT: O mapa. / EN-UK: The map.
    :param path:
        PT-PT: Destino. A pasta é criada se faltar.
        EN-UK: Destination. The folder is created when missing.
    :param started:
        PT-PT: Quando o mapeamento começou, para o resumo.
        EN-UK: When the mapping started, for the summary.
    :return:
        PT-PT: O caminho gravado. / EN-UK: The written path.
    :raises ExcelError:
        PT-PT: Se o openpyxl não estiver instalado.
        EN-UK: If openpyxl is not installed.
    """
    try:
        from openpyxl import Workbook
    except ImportError as exc:
        raise ExcelError("O openpyxl não está instalado. Instale com: pip install openpyxl") from exc

    livro = Workbook()
    livro.remove(livro.active)

    _sheet_summary(livro, topology, started or datetime.now())
    _sheet_devices(livro, topology)
    _sheet_links(livro, topology)
    _sheet_endpoints(livro, topology)
    _sheet_issues(livro, topology)

    path.parent.mkdir(parents=True, exist_ok=True)
    livro.save(path)
    return path


# ---------------------------------------------------------------------------
# PT-PT: As folhas.
# EN-UK: The sheets.
# ---------------------------------------------------------------------------


def _sheet_summary(workbook: Any, topology: Topology, started: datetime) -> None:
    """PT-PT: O que correu. / EN-UK: What ran."""
    folha = workbook.create_sheet("Resumo")
    localizados = sum(1 for ponto in topology.endpoints if ponto.located)
    sem_fios = sum(1 for ponto in topology.endpoints if ponto.wireless)

    linhas = [
        ("Ferramenta", f"{__app_name__} {__version__}"),
        ("Data do mapeamento", started.strftime("%Y-%m-%d %H:%M")),
        ("", ""),
        ("Equipamentos alcançados", len(topology.reached)),
        ("Equipamentos por alcançar", len(topology.unreached)),
        ("Ligações entre equipamentos", len(topology.links)),
        ("", ""),
        ("Pontos finais encontrados", len(topology.endpoints)),
        ("Com porta identificada", localizados),
        ("Sem fios", sem_fios),
        ("", ""),
        ("Problemas assinalados", len(topology.issues)),
    ]

    for confianca in Confidence:
        quantos = sum(1 for ponto in topology.endpoints if ponto.confidence is confianca)
        linhas.append((f"Classificação de confiança {confianca.value.lower()}", quantos))

    folha.append(["Resumo do mapeamento", ""])
    _style_title(folha, 1)
    for rotulo, valor in linhas:
        folha.append([rotulo, valor])

    _widths(folha, [38, 46])
    _note(
        folha,
        len(linhas) + 3,
        "As classificações de confiança baixa ou nenhuma não estão erradas — estão por "
        "confirmar. Veja a coluna dos sinais na folha dos pontos finais para saber em que "
        "é que cada uma se baseou.",
    )


def _sheet_devices(workbook: Any, topology: Topology) -> None:
    """PT-PT: Os switches. / EN-UK: The switches."""
    folha = workbook.create_sheet("Equipamentos")
    cabecalho = [
        "Nome",
        "Endereço",
        "Plataforma",
        "Modelo",
        "Alcançado",
        "Saltos",
        "Descoberto por",
        "Portas lidas",
        "Endereços MAC",
        "Vizinhos",
        "Linhas não interpretadas",
        "Erro",
    ]
    folha.append(cabecalho)
    _style_header(folha)

    for dispositivo in sorted(topology.devices.values(), key=lambda d: d.label.lower()):
        folha.append(
            [
                dispositivo.label,
                dispositivo.host,
                dispositivo.platform.label,
                dispositivo.model,
                "sim" if dispositivo.reached else "NÃO",
                dispositivo.depth,
                dispositivo.source.value,
                len(dispositivo.facts.ports),
                len(dispositivo.facts.macs),
                len(dispositivo.facts.neighbours),
                dispositivo.facts.unparsed_lines,
                dispositivo.error,
            ]
        )

    _widths(folha, [24, 16, 22, 26, 11, 8, 16, 13, 15, 11, 22, 50])
    _finish(folha, len(cabecalho))


def _sheet_links(workbook: Any, topology: Topology) -> None:
    """PT-PT: Os cabos entre switches. / EN-UK: The cables between switches."""
    folha = workbook.create_sheet("Ligações")
    cabecalho = ["Equipamento A", "Porta A", "Equipamento B", "Porta B", "Descoberta por"]
    folha.append(cabecalho)
    _style_header(folha)

    for ligacao in topology.links:
        folha.append(
            [
                ligacao.a_device,
                ligacao.a_port,
                ligacao.b_device,
                ligacao.b_port,
                ligacao.source.value,
            ]
        )

    _widths(folha, [24, 20, 24, 20, 18])
    _finish(folha, len(cabecalho))


def _sheet_endpoints(workbook: Any, topology: Topology) -> None:
    """PT-PT: A folha grande. / EN-UK: The big sheet."""
    from openpyxl.styles import PatternFill

    folha = workbook.create_sheet("Pontos finais")
    cabecalho = [
        "Equipamento",
        "Porta",
        "Etiqueta da porta",
        "VLAN",
        "Tipo",
        "Confiança",
        "Endereço MAC",
        "Endereço IP",
        "Nome",
        "Fabricante",
        "PoE (W)",
        "Sem fios",
        "Ponto de acesso",
        "Notas",
        "Sinais que sustentam a classificação",
    ]
    folha.append(cabecalho)
    _style_header(folha)

    ordenados = sorted(
        topology.endpoints,
        key=lambda ponto: (
            ponto.switch.lower(),
            _port_sort_key(ponto.port),
            ponto.mac,
        ),
    )

    for ponto in ordenados:
        folha.append(
            [
                ponto.switch,
                ponto.port,
                ponto.port_description,
                ponto.vlan,
                ponto.role.value,
                ponto.confidence.value,
                ponto.mac,
                ponto.ip,
                ponto.hostname,
                ponto.vendor,
                ponto.poe_watts,
                "sim" if ponto.wireless else "",
                ponto.access_point,
                ponto.note,
                " | ".join(ponto.signals),
            ]
        )
        cor = _CONFIDENCE_FILL[ponto.confidence]
        folha.cell(row=folha.max_row, column=6).fill = PatternFill("solid", fgColor=cor)

    _widths(folha, [22, 18, 24, 7, 20, 11, 20, 16, 26, 24, 9, 9, 20, 46, 70])
    _finish(folha, len(cabecalho))


def _sheet_issues(workbook: Any, topology: Topology) -> None:
    """PT-PT: O que vale a pena olhar. / EN-UK: What is worth a look."""
    folha = workbook.create_sheet("Problemas")
    cabecalho = ["Gravidade", "Onde", "O que se passa"]
    folha.append(cabecalho)
    _style_header(folha)

    ordem = {"ERRO": 0, "AVISO": 1, "INFO": 2}
    for problema in sorted(topology.issues, key=lambda i: (ordem.get(i.severity, 3), i.subject)):
        folha.append([problema.severity, problema.subject, problema.message])

    _widths(folha, [12, 30, 110])
    _finish(folha, len(cabecalho))


# ---------------------------------------------------------------------------
# PT-PT: Apresentação.
# EN-UK: Presentation.
# ---------------------------------------------------------------------------


def _style_header(sheet: Any) -> None:
    """PT-PT: Cabeçalho a negrito sobre fundo escuro. / EN-UK: Bold header on a dark fill."""
    from openpyxl.styles import Font, PatternFill

    for celula in sheet[1]:
        celula.font = Font(bold=True, color="FFFFFF")
        celula.fill = PatternFill("solid", fgColor=_HEADER_FILL)


def _style_title(sheet: Any, row: int) -> None:
    """PT-PT: Título da folha de resumo. / EN-UK: The summary sheet's title."""
    from openpyxl.styles import Font

    sheet.cell(row=row, column=1).font = Font(bold=True, size=14)


def _widths(sheet: Any, widths: list[int]) -> None:
    """PT-PT: Larguras de coluna. / EN-UK: Column widths."""
    for indice, largura in enumerate(widths, start=1):
        sheet.column_dimensions[sheet.cell(row=1, column=indice).column_letter].width = largura


def _finish(sheet: Any, columns: int) -> None:
    """
    PT-PT: Fixa o cabeçalho e liga o filtro automático.

           Sem isto, uma folha com dois mil pontos finais é inutilizável: rola-se
           dez linhas e já não se sabe que coluna é qual.

    EN-UK: Freezes the header and switches on the auto-filter.

           Without this, a sheet with two thousand endpoints is unusable: scroll
           ten rows and you no longer know which column is which.
    """
    sheet.freeze_panes = "A2"
    if sheet.max_row > 1:
        ultima = sheet.cell(row=1, column=columns).column_letter
        sheet.auto_filter.ref = f"A1:{ultima}{sheet.max_row}"


def _note(sheet: Any, row: int, text: str) -> None:
    """PT-PT: Uma nota em texto pequeno. / EN-UK: A note in small text."""
    from openpyxl.styles import Alignment, Font

    celula = sheet.cell(row=row, column=1)
    celula.value = text
    celula.font = Font(size=9, italic=True)
    celula.alignment = Alignment(wrap_text=True, vertical="top")
    sheet.merge_cells(start_row=row, start_column=1, end_row=row + 3, end_column=2)


def _port_sort_key(port: str) -> tuple[int, ...]:
    """
    PT-PT: Ordena portas por número e não por texto.

           Alfabeticamente, a porta 10 vem antes da 2, e uma folha de 48 portas
           fica com a ordem trocada de uma ponta à outra. Aqui separa-se o nome
           nos seus números e ordena-se por eles.

    EN-UK: Sorts ports numerically rather than as text.

           Alphabetically, port 10 comes before port 2, and a 48-port sheet ends
           up jumbled from end to end. Here the name is split into its numbers
           and ordered by them.
    """
    import re

    numeros = tuple(int(parte) for parte in re.findall(r"\d+", port))
    return numeros or (0,)
