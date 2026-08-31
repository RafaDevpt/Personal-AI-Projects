#!/usr/bin/env python3
"""
PT-PT: Testes da classificação, dos fabricantes, da garantia de só-leitura e
       dos relatórios.

       A classe que mais importa aqui é a `TestSoLeitura`. A garantia de que
       este programa não escreve em equipamento nenhum não pode depender de
       ninguém se lembrar: tem de estar testada, e tem de falhar ruidosamente se
       alguém um dia acrescentar um comando que escreve.

EN-UK: Tests for the classification, the manufacturers, the read-only guarantee
       and the reports.

       The class that matters most here is `TestSoLeitura`. The guarantee that
       this program writes to no device cannot depend on anyone remembering: it
       has to be tested, and it has to fail loudly if somebody one day adds a
       command that writes.

Created by Redfox using Claude
"""

from __future__ import annotations

from pathlib import Path

import pytest

from netmap import classify, oui
from netmap.classify import PortContext
from netmap.collector import UnsafeCommandError, assert_read_only, is_read_only
from netmap.models import (
    Confidence,
    Endpoint,
    LldpNeighbour,
    Role,
    Source,
    is_locally_administered,
    normalise_mac,
    normalise_port,
)
from netmap.parsers import supported_platforms

# ---------------------------------------------------------------------------
# PT-PT: Normalização — a base de todo o cruzamento.
# ---------------------------------------------------------------------------


class TestNormalizacao:
    """PT-PT: Sem isto, nada cruza. / EN-UK: Without this, nothing cross-references."""

    @pytest.mark.parametrize(
        "escrita",
        [
            "aa:bb:cc:dd:ee:ff",
            "AA:BB:CC:DD:EE:FF",
            "aabb.ccdd.eeff",
            "aabb-ccdd-eeff",
            "AABB-CCDD-EEFF",
            "aabbccddeeff",
            "  aa:bb:cc:dd:ee:ff  ",
        ],
    )
    def test_todas_as_escritas_dao_o_mesmo(self, escrita: str) -> None:
        # PT-PT: O Cisco, o Aruba e o UniFi escrevem o mesmo MAC de três
        #        maneiras. Se não coincidirem, o mapa sai vazio sem dar erro.
        # EN-UK: Cisco, Aruba and UniFi write the same MAC three ways. If they
        #        do not match, the map comes out empty with no error.
        assert normalise_mac(escrita) == "aa:bb:cc:dd:ee:ff"

    @pytest.mark.parametrize("lixo", ["", "não é um mac", "aa:bb:cc", "12345"])
    def test_texto_que_nao_e_mac(self, lixo: str) -> None:
        assert normalise_mac(lixo) == ""

    @pytest.mark.parametrize(
        ("escrita", "esperado"),
        [
            ("GigabitEthernet1/0/1", "Gi1/0/1"),
            ("gigabitethernet1/0/1", "Gi1/0/1"),
            ("Gi1/0/1", "Gi1/0/1"),
            ("TenGigabitEthernet1/1/1", "Te1/1/1"),
            ("FastEthernet0/1", "Fa0/1"),
            ("1/1/48", "1/1/48"),
            ("0/24", "0/24"),
        ],
    )
    def test_nomes_de_porta(self, escrita: str, esperado: str) -> None:
        assert normalise_port(escrita) == esperado

    def test_mac_administrado_localmente(self) -> None:
        # PT-PT: É o que os telemóveis modernos apresentam por omissão.
        # EN-UK: It is what modern phones present by default.
        assert is_locally_administered("02:11:22:33:44:55")
        assert not is_locally_administered("00:50:56:11:22:33")


# ---------------------------------------------------------------------------
# PT-PT: Fabricantes.
# ---------------------------------------------------------------------------


class TestFabricantes:
    """PT-PT: O OUI e o que se conclui dele. / EN-UK: The OUI and what follows from it."""

    def test_vmware(self) -> None:
        assert oui.lookup("00:50:56:aa:bb:cc") == "VMware"

    def test_familia_de_virtualizacao(self) -> None:
        assert oui.family("VMware") == "virtual"
        assert oui.family("Microsoft Hyper-V") == "virtual"

    def test_oui_desconhecido_devolve_vazio(self) -> None:
        # PT-PT: Vazio e não "desconhecido": vazio é ausência de sinal, e a
        #        classificação sabe o que fazer com isso.
        #
        #        O `9c` do início não é arbitrário: um MAC de teste tem de ser
        #        globalmente administrado, senão a resposta é "MAC aleatório" e
        #        não se chega a testar a procura. Endereços como `aa:` ou `ab:`
        #        têm o bit de administração local ligado.
        #
        # EN-UK: Empty rather than "unknown": empty is absence of signal, and
        #        the classifier knows what to do with that.
        #
        #        The leading `9c` is not arbitrary: a test MAC has to be
        #        globally administered, otherwise the answer is "random MAC" and
        #        the lookup never gets tested. Addresses like `aa:` or `ab:`
        #        have the locally-administered bit set.
        assert oui.lookup("9c:bb:cc:dd:ee:ff") == ""

    def test_mac_aleatorio_e_informacao(self) -> None:
        assert "aleatório" in oui.lookup("02:11:22:33:44:55")

    def test_importar_csv_do_ieee(self, tmp_path: Path) -> None:
        ficheiro = tmp_path / "oui.csv"
        ficheiro.write_text(
            "Registry,Assignment,Organization Name,Organization Address\n"
            "MA-L,9CCDEF,Fabricante Inventado Lda,Rua Qualquer\n",
            encoding="utf-8",
        )
        try:
            assert oui.import_ieee_file(ficheiro) == 1
            assert oui.lookup("9c:cd:ef:11:22:33") == "Fabricante Inventado Lda"
        finally:
            oui.clear_imported()

    def test_importar_txt_do_ieee(self, tmp_path: Path) -> None:
        ficheiro = tmp_path / "oui.txt"
        ficheiro.write_text(
            "9C-CD-EF   (hex)\t\tFabricante Inventado Lda\n", encoding="utf-8"
        )
        try:
            assert oui.import_ieee_file(ficheiro) == 1
            assert oui.lookup("9c:cd:ef:11:22:33") == "Fabricante Inventado Lda"
        finally:
            oui.clear_imported()

    def test_a_tabela_curada_nao_esta_vazia(self) -> None:
        assert oui.curated_count() > 40


# ---------------------------------------------------------------------------
# PT-PT: Classificação.
# ---------------------------------------------------------------------------


def ponto(**campos: object) -> Endpoint:
    """PT-PT: Um ponto final para testar. / EN-UK: An endpoint to test with."""
    base: dict = {"mac": "aa:bb:cc:dd:ee:ff"}
    base.update(campos)
    return Endpoint(**base)  # type: ignore[arg-type]


def vizinho(*capacidades: str) -> LldpNeighbour:
    """PT-PT: Um vizinho com estas capacidades. / EN-UK: A neighbour with these capabilities."""
    return LldpNeighbour(
        local_port="Gi1/0/1", remote_name="algo", capabilities=set(capacidades), source=Source.LLDP
    )


class TestClassificacaoPorLldp:
    """PT-PT: O único sinal em que o equipamento fala por si."""

    def test_ap(self) -> None:
        resultado = classify.classify(ponto(), PortContext(neighbour=vizinho("wlan-ap")))
        assert resultado.role is Role.ACCESS_POINT
        assert resultado.confidence is Confidence.HIGH

    def test_telefone(self) -> None:
        resultado = classify.classify(ponto(), PortContext(neighbour=vizinho("bridge", "telephone")))
        assert resultado.role is Role.PHONE

    def test_switch(self) -> None:
        resultado = classify.classify(ponto(), PortContext(neighbour=vizinho("bridge")))
        assert resultado.role is Role.SWITCH

    def test_o_lldp_ganha_ao_oui(self) -> None:
        # PT-PT: Um AP com placa Intel é um AP. O que o equipamento diz de si
        #        vale mais do que o fabricante do chip de rede.
        # EN-UK: An AP with an Intel NIC is an AP. What the device says about
        #        itself beats the maker of its network chip.
        resultado = classify.classify(
            ponto(vendor="Intel"), PortContext(neighbour=vizinho("wlan-ap"))
        )
        assert resultado.role is Role.ACCESS_POINT


class TestClassificacaoPorNome:
    """PT-PT: Os nomes que vêm de fábrica. / EN-UK: The factory-assigned names."""

    def test_impressora_hp_pelo_nome_npi(self) -> None:
        # PT-PT: Ninguém escolhe um nome `NPI1A2B3C` — vem da JetDirect.
        # EN-UK: Nobody chooses a name like `NPI1A2B3C` — it comes from JetDirect.
        resultado = classify.classify(ponto(hostname="NPI1A2B3C"))
        assert resultado.role is Role.PRINTER
        assert resultado.confidence is Confidence.HIGH

    def test_impressora_brother(self) -> None:
        assert classify.classify(ponto(hostname="BRN30055C112233")).role is Role.PRINTER

    def test_nome_de_posto(self) -> None:
        assert classify.classify(ponto(hostname="DESKTOP-A1B2C3")).role is Role.PC


class TestClassificacaoPorFabricante:
    """PT-PT: O que o OUI sugere, e o que não sugere."""

    def test_maquina_virtual(self) -> None:
        resultado = classify.classify(ponto(vendor="VMware"))
        assert resultado.role is Role.VIRTUAL
        assert resultado.confidence is Confidence.HIGH

    def test_hp_e_um_empate_declarado(self) -> None:
        # PT-PT: A HP fabrica postos e impressoras com o mesmo OUI. Escolher um
        #        seria acertar metade das vezes e enganar a outra metade.
        # EN-UK: HP makes workstations and printers under the same OUI. Picking
        #        one would be right half the time and misleading the other half.
        resultado = classify.classify(ponto(vendor="Hewlett-Packard"))
        assert resultado.confidence is Confidence.NONE
        assert "conflito" in resultado.note.lower()

    def test_sem_sinais_nenhuns(self) -> None:
        resultado = classify.classify(ponto())
        assert resultado.role is Role.UNKNOWN
        assert resultado.confidence is Confidence.NONE
        assert resultado.signals


class TestClassificacaoPorPoe:
    """PT-PT: O consumo confirma ou desmente. / EN-UK: Draw confirms or contradicts."""

    def test_consumo_alto_sugere_ap(self) -> None:
        resultado = classify.classify(ponto(), PortContext(poe_watts=14.2))
        assert resultado.role is Role.ACCESS_POINT

    def test_consumo_baixo_sugere_telefone(self) -> None:
        resultado = classify.classify(ponto(), PortContext(poe_watts=5.5))
        assert resultado.role is Role.PHONE

    def test_sem_consumo_nao_diz_nada(self) -> None:
        resultado = classify.classify(ponto(), PortContext(poe_watts=0.0))
        assert resultado.role is Role.UNKNOWN


class TestSinaisEmConflito:
    """PT-PT: Duas certezas contraditórias não fazem uma certeza."""

    def test_a_confianca_desce_e_diz_se_porque(self) -> None:
        resultado = classify.classify(
            ponto(hostname="NPI1A2B3C"), PortContext(neighbour=vizinho("wlan-ap"))
        )
        assert resultado.confidence is Confidence.MEDIUM
        assert "conflito" in resultado.note.lower()

    def test_todos_os_sinais_ficam_registados(self) -> None:
        resultado = classify.classify(
            ponto(vendor="VMware", hostname="DESKTOP-A1"), PortContext(poe_watts=12.0)
        )
        assert len(resultado.signals) >= 3


class TestSwitchNaoGerido:
    """PT-PT: Uma conclusão sobre a porta, não sobre o equipamento."""

    def test_muitos_macs_sem_vizinho(self) -> None:
        assert classify.unmanaged_switch_suspected(PortContext(macs_on_port=6))

    def test_dois_macs_sao_normais(self) -> None:
        # PT-PT: Um telefone com um posto atrás. Não é um switch escondido.
        # EN-UK: A phone with a workstation behind it. Not a hidden switch.
        assert not classify.unmanaged_switch_suspected(PortContext(macs_on_port=2))

    def test_com_vizinho_lldp_nao_conta(self) -> None:
        assert not classify.unmanaged_switch_suspected(
            PortContext(macs_on_port=20, neighbour=vizinho("bridge"))
        )

    def test_sem_fios_nao_conta(self) -> None:
        assert not classify.unmanaged_switch_suspected(
            PortContext(macs_on_port=30, wireless=True)
        )


# ---------------------------------------------------------------------------
# PT-PT: A garantia de só-leitura.
# ---------------------------------------------------------------------------


class TestSoLeitura:
    """
    PT-PT: A propriedade que este programa promete e tem de cumprir.
    EN-UK: The property this program promises and must keep.
    """

    @pytest.mark.parametrize(
        "comando",
        [
            "show running-config",
            "show mac address-table",
            "display lldp neighbor",
            "telnet localhost",
        ],
    )
    def test_comandos_de_leitura_passam(self, comando: str) -> None:
        assert is_read_only(comando)

    @pytest.mark.parametrize(
        "comando",
        [
            "configure terminal",
            "write memory",
            "reload",
            "no vlan 10",
            "copy running-config startup-config",
            "",
        ],
    )
    def test_comandos_que_escrevem_sao_recusados(self, comando: str) -> None:
        assert not is_read_only(comando)

    @pytest.mark.parametrize(
        "comando",
        ["show version; reload", "show version && write memory", "show run || reload"],
    )
    def test_encadeamento_e_recusado(self, comando: str) -> None:
        # PT-PT: Um `show` seguido de um `;` já não é um `show`.
        # EN-UK: A `show` followed by a `;` is no longer a `show`.
        assert not is_read_only(comando)

    def test_todos_os_comandos_de_todos_os_leitores(self) -> None:
        # PT-PT: Este é o teste que interessa: se alguém acrescentar um comando
        #        que escreve a qualquer plataforma, falha aqui.
        # EN-UK: This is the test that matters: if anyone adds a writing command
        #        to any platform, it fails here.
        from netmap.parsers import get_parser

        for plataforma in supported_platforms():
            assert_read_only(get_parser(plataforma).commands)

    def test_um_comando_que_escreve_rebenta(self) -> None:
        with pytest.raises(UnsafeCommandError, match="não é de leitura"):
            assert_read_only({"mau": "reload"})


# ---------------------------------------------------------------------------
# PT-PT: Relatórios.
# ---------------------------------------------------------------------------


class TestRelatorios:
    """PT-PT: Que se escrevem e que levam o que interessa."""

    @pytest.fixture
    def mapa(self):
        import sys

        sys.path.insert(0, str(Path(__file__).parent))
        from netmap import topology as topo
        from netmap.crawler import CrawlOptions, crawl
        from netmap.models import NetworkDevice, Platform
        from test_crawl_topology import CREDENCIAIS, fake_collect

        semente = NetworkDevice(host="10.0.10.1", name="SW-CORE-01", platform=Platform.CISCO_IOS)
        resultado = crawl([semente], CREDENCIAIS, CrawlOptions(), collect_fn=fake_collect)
        mapa = topo.build(resultado.devices)
        mapa.issues = resultado.issues + mapa.issues
        return mapa

    def test_excel_escreve_as_cinco_folhas(self, mapa, tmp_path: Path) -> None:
        from openpyxl import load_workbook

        from netmap.reports import write_excel

        caminho = write_excel(mapa, tmp_path / "mapa.xlsx")
        livro = load_workbook(caminho)
        assert livro.sheetnames == [
            "Resumo",
            "Equipamentos",
            "Ligações",
            "Pontos finais",
            "Problemas",
        ]

    def test_o_excel_leva_todos_os_pontos_finais(self, mapa, tmp_path: Path) -> None:
        from openpyxl import load_workbook

        from netmap.reports import write_excel

        caminho = write_excel(mapa, tmp_path / "mapa.xlsx")
        folha = load_workbook(caminho)["Pontos finais"]
        assert folha.max_row == len(mapa.endpoints) + 1

    def test_o_excel_leva_os_sinais(self, mapa, tmp_path: Path) -> None:
        # PT-PT: A coluna que separa esta ferramenta de uma que adivinha.
        # EN-UK: The column that separates this tool from one that guesses.
        from openpyxl import load_workbook

        from netmap.reports import write_excel

        caminho = write_excel(mapa, tmp_path / "mapa.xlsx")
        folha = load_workbook(caminho)["Pontos finais"]
        cabecalhos = [celula.value for celula in folha[1]]
        assert "Sinais que sustentam a classificação" in cabecalhos

    def test_pdf_escreve(self, mapa, tmp_path: Path) -> None:
        from netmap.reports import write_pdf

        caminho = write_pdf(mapa, tmp_path / "mapa.pdf")
        assert caminho.exists()
        assert caminho.read_bytes().startswith(b"%PDF")

    def test_pdf_com_um_mapa_vazio(self, tmp_path: Path) -> None:
        # PT-PT: Um mapeamento que não alcançou nada tem de produzir um
        #        relatório que o diga, e não um traceback.
        # EN-UK: A mapping that reached nothing must produce a report saying so,
        #        rather than a traceback.
        from netmap.models import Topology
        from netmap.reports import write_pdf

        caminho = write_pdf(Topology(), tmp_path / "vazio.pdf")
        assert caminho.exists()
