#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
PT-PT: Inventário de impressoras em Excel.

       Esta é a mudança central face à versão anterior, que trazia as 24
       impressoras do Penha Longa escritas dentro do código. Isso tornava a
       ferramenta impossível de partilhar (o inventário é informação interna) e
       obrigava a editar Python para acrescentar uma impressora.

       Agora o inventário vive num ficheiro .xlsx que qualquer pessoa edita, e
       que a aplicação cria sozinha na primeira execução, já com uma linha de
       exemplo e uma folha de instruções.

EN-UK: Printer inventory in Excel.

       This is the central change from the previous version, which carried the
       24 Penha Longa printers written inside the code. That made the tool
       impossible to share (the inventory is internal information) and forced
       anyone adding a printer to edit Python.

       The inventory now lives in an .xlsx file that anybody can edit, and which
       the application creates by itself on first run, complete with an example
       row and an instructions sheet.

PT-PT: Porquê Excel e não CSV. As pessoas que mantêm a lista de impressoras
       abrem-na no Excel de qualquer forma, e o Excel estraga CSV com acentos
       silenciosamente ao gravar. O .xlsx guarda o texto em UTF-8 sem margem
       para dúvida. Mesmo assim, ler CSV continua suportado para quem já tenha
       uma lista nesse formato.

EN-UK: Why Excel rather than CSV. The people who maintain the printer list open
       it in Excel anyway, and Excel silently mangles accented CSV on save.
       .xlsx stores the text as UTF-8 with no room for doubt. CSV reading is
       still supported, for anyone who already has a list in that format.

Created by Redfox using Claude
"""

from __future__ import annotations

import csv
import ipaddress
import logging
from pathlib import Path

from .models import Printer

_log = logging.getLogger(__name__)

# PT-PT: Nome da folha de dados. Fixo de propósito: se o utilizador lhe mudar o
#        nome, a aplicação recorre à primeira folha do livro.
# EN-UK: Name of the data sheet. Deliberately fixed: if the user renames it, the
#        application falls back to the workbook's first sheet.
SHEET_NAME = "Impressoras"
INSTRUCTIONS_SHEET = "Instruções"

# PT-PT: Cabeçalhos esperados, na ordem em que são escritos. A leitura procura
#        as colunas pelo nome, e não pela posição, para que reordenar colunas no
#        Excel não parta o ficheiro.
# EN-UK: Expected headers, in the order they are written. Reading looks the
#        columns up by name rather than by position, so reordering columns in
#        Excel does not break the file.
COLUMNS: tuple[str, ...] = (
    "IP",
    "Localização",
    "Nome de rede",
    "Modelo",
    "Número de série",
    "MAC",
    "Protocolo",
    "Activa",
    "Notas",
)

# PT-PT: Nomes alternativos aceites na leitura, para tolerar ficheiros feitos à
#        mão ou exportados de outra ferramenta.
# EN-UK: Alternative names accepted when reading, to tolerate hand-made files or
#        exports from another tool.
_HEADER_ALIASES: dict[str, str] = {
    "ip": "IP",
    "endereço": "IP",
    "endereco": "IP",
    "address": "IP",
    "localização": "Localização",
    "localizacao": "Localização",
    "local": "Localização",
    "location": "Localização",
    "nome de rede": "Nome de rede",
    "hostname": "Nome de rede",
    "nome": "Nome de rede",
    "modelo": "Modelo",
    "model": "Modelo",
    "número de série": "Número de série",
    "numero de serie": "Número de série",
    "s/n": "Número de série",
    "sn": "Número de série",
    "serial": "Número de série",
    "mac": "MAC",
    "protocolo": "Protocolo",
    "scheme": "Protocolo",
    "activa": "Activa",
    "ativa": "Activa",
    "enabled": "Activa",
    "notas": "Notas",
    "notes": "Notas",
}

# PT-PT: Valores aceites como "sim" na coluna Activa. Aceitar várias formas
#        evita que a lista deixe de funcionar porque alguém escreveu "SIM".
# EN-UK: Values accepted as "yes" in the Activa column. Accepting several forms
#        stops the list breaking because somebody typed "SIM".
_TRUTHY = {"sim", "s", "yes", "y", "true", "1", "verdadeiro", "x", "activa", "ativa"}


class InventoryError(RuntimeError):
    """
    PT-PT: Erro de inventário com mensagem destinada ao utilizador final.
    EN-UK: Inventory error carrying a message intended for the end user.
    """


# ---------------------------------------------------------------------------
# PT-PT: Leitura / EN-UK: Reading
# ---------------------------------------------------------------------------


def _normalise_header(value: object) -> str:
    """
    PT-PT: Converte um cabeçalho lido para o nome canónico da coluna.
    EN-UK: Converts a header as read into the column's canonical name.

    :param value:
        PT-PT: Conteúdo da célula de cabeçalho.
        EN-UK: Contents of the header cell.
    :return:
        PT-PT: Nome canónico, ou string vazia se não for reconhecido.
        EN-UK: Canonical name, or an empty string if unrecognised.
    """
    if value is None:
        return ""
    text = str(value).strip().lower()
    return _HEADER_ALIASES.get(text, "")


def _row_to_printer(row: dict[str, str]) -> Printer | None:
    """
    PT-PT: Converte uma linha lida num objecto Printer, validando o IP.

           Uma linha sem IP válido é ignorada em silêncio ao nível do
           utilizador (fica registada no log). Linhas em branco no fim da folha
           são normais num ficheiro editado à mão e não devem gerar avisos.

    EN-UK: Converts a read row into a Printer object, validating the IP.

           A row with no valid IP is ignored silently from the user's point of
           view (it is recorded in the log). Blank rows at the end of the sheet
           are normal in a hand-edited file and should not raise warnings.

    :param row:
        PT-PT: Mapa coluna -> valor. / EN-UK: Column -> value mapping.
    :return:
        PT-PT: Impressora, ou None se a linha não for utilizável.
        EN-UK: Printer, or None if the row is unusable.
    """
    raw_ip = (row.get("IP") or "").strip()
    if not raw_ip:
        return None

    try:
        # PT-PT: Validar aqui evita que um erro de escrita só apareça mais
        #        tarde como um timeout de rede inexplicável.
        # EN-UK: Validating here stops a typo surfacing later as an
        #        unexplained network timeout.
        ipaddress.ip_address(raw_ip)
    except ValueError:
        _log.warning("IP inválido ignorado: %r", raw_ip)
        return None

    scheme = (row.get("Protocolo") or "http").strip().lower()
    if scheme not in ("http", "https"):
        scheme = "http"

    activa = (row.get("Activa") or "sim").strip().lower()

    return Printer(
        ip=raw_ip,
        location=(row.get("Localização") or "").strip(),
        hostname=(row.get("Nome de rede") or "").strip(),
        model=(row.get("Modelo") or "").strip(),
        serial=(row.get("Número de série") or "").strip(),
        mac=(row.get("MAC") or "").strip(),
        scheme=scheme,
        enabled=activa in _TRUTHY,
        notes=(row.get("Notas") or "").strip(),
    )


def load_xlsx(path: Path) -> list[Printer]:
    """
    PT-PT: Lê o inventário de um ficheiro Excel.

    EN-UK: Reads the inventory from an Excel file.

    :param path:
        PT-PT: Caminho do ficheiro .xlsx. / EN-UK: Path of the .xlsx file.
    :raises InventoryError:
        PT-PT: Se o openpyxl não estiver instalado ou o ficheiro for ilegível.
        EN-UK: If openpyxl is not installed or the file cannot be read.
    """
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise InventoryError(
            "Falta a biblioteca openpyxl, necessária para ler ficheiros Excel.\n"
            "Execute: pip install -r requirements.txt"
        ) from exc

    try:
        # PT-PT: read_only acelera muito em ficheiros grandes; data_only
        #        devolve o resultado das fórmulas em vez do seu texto.
        # EN-UK: read_only is much faster on large files; data_only returns
        #        formula results rather than the formula text.
        workbook = load_workbook(path, read_only=True, data_only=True)
    except (OSError, ValueError, KeyError) as exc:
        raise InventoryError(
            f"Não foi possível abrir {path.name}.\n"
            f"Confirme que é um ficheiro Excel válido e que não está aberto "
            f"noutro programa.\nDetalhe / detail: {exc}"
        ) from exc

    try:
        sheet = (
            workbook[SHEET_NAME]
            if SHEET_NAME in workbook.sheetnames
            else workbook[workbook.sheetnames[0]]
        )

        rows = sheet.iter_rows(values_only=True)
        try:
            header_row = next(rows)
        except StopIteration:
            _log.warning("Folha vazia em %s", path)
            return []

        # PT-PT: Índice de coluna -> nome canónico. Colunas desconhecidas são
        #        simplesmente ignoradas, o que permite ao utilizador acrescentar
        #        colunas suas (custo, data de instalação) sem quebrar nada.
        # EN-UK: Column index -> canonical name. Unknown columns are simply
        #        ignored, which lets the user add their own columns (cost,
        #        install date) without breaking anything.
        mapping: dict[int, str] = {}
        for index, cell in enumerate(header_row):
            canonical = _normalise_header(cell)
            if canonical:
                mapping[index] = canonical

        if "IP" not in mapping.values():
            raise InventoryError(
                f"O ficheiro {path.name} não tem uma coluna 'IP'.\n"
                f"Colunas esperadas: {', '.join(COLUMNS)}"
            )

        printers: list[Printer] = []
        for row in rows:
            record = {
                mapping[index]: ("" if row[index] is None else str(row[index]))
                for index in mapping
                if index < len(row)
            }
            printer = _row_to_printer(record)
            if printer is not None:
                printers.append(printer)

    finally:
        # PT-PT: Em modo read_only o ficheiro fica aberto até se fechar
        #        explicitamente, e em Windows isso impede o utilizador de o
        #        abrir no Excel.
        # EN-UK: In read_only mode the file stays open until explicitly closed,
        #        and on Windows that stops the user opening it in Excel.
        workbook.close()

    _log.info("Inventário: %d impressoras lidas de %s", len(printers), path.name)
    return printers


def load_csv(path: Path) -> list[Printer]:
    """
    PT-PT: Lê o inventário de um CSV, para quem já tenha a lista nesse formato.
           Detecta o separador automaticamente: o Excel português grava com
           ponto e vírgula, e uma lista feita à mão costuma usar vírgula.

    EN-UK: Reads the inventory from a CSV, for anyone who already has the list
           in that format. The separator is detected automatically: Portuguese
           Excel saves with semicolons, and a hand-made list usually uses commas.

    :param path:
        PT-PT: Caminho do ficheiro .csv. / EN-UK: Path of the .csv file.
    """
    try:
        # PT-PT: utf-8-sig remove a marca de ordem de bytes que o Excel escreve
        #        e que, sem isto, contaminaria o primeiro cabeçalho.
        # EN-UK: utf-8-sig strips the byte-order mark Excel writes, which would
        #        otherwise contaminate the first header.
        text = path.read_text(encoding="utf-8-sig")
    except (OSError, UnicodeDecodeError) as exc:
        raise InventoryError(f"Não foi possível ler {path.name}: {exc}") from exc

    sample = text[:2048]
    delimiter = ";" if sample.count(";") > sample.count(",") else ","

    printers: list[Printer] = []
    for raw in csv.DictReader(text.splitlines(), delimiter=delimiter):
        record = {
            _normalise_header(key): (value or "")
            for key, value in raw.items()
            if _normalise_header(key)
        }
        printer = _row_to_printer(record)
        if printer is not None:
            printers.append(printer)

    _log.info("Inventário: %d impressoras lidas de %s", len(printers), path.name)
    return printers


def load(path: Path) -> list[Printer]:
    """
    PT-PT: Lê o inventário, escolhendo o leitor pela extensão do ficheiro.

    EN-UK: Reads the inventory, choosing the reader by the file extension.

    :param path:
        PT-PT: Caminho do inventário. / EN-UK: Path of the inventory.
    :raises InventoryError:
        PT-PT: Se o ficheiro não existir ou o formato não for suportado.
        EN-UK: If the file does not exist or the format is unsupported.
    """
    if not path.is_file():
        raise InventoryError(
            f"Inventário não encontrado: {path}\n"
            f"Use 'Criar modelo Excel' para gerar um ficheiro novo."
        )

    suffix = path.suffix.lower()
    if suffix in (".xlsx", ".xlsm"):
        return load_xlsx(path)
    if suffix == ".csv":
        return load_csv(path)

    raise InventoryError(
        f"Formato não suportado: {suffix}\nUse .xlsx ou .csv."
    )


# ---------------------------------------------------------------------------
# PT-PT: Escrita / EN-UK: Writing
# ---------------------------------------------------------------------------


def save_xlsx(printers: list[Printer], path: Path) -> Path:
    """
    PT-PT: Grava a lista de impressoras num ficheiro Excel formatado.
           Usado quando a descoberta na rede encontra equipamento novo e o
           utilizador quer guardá-lo no inventário.

    EN-UK: Writes the printer list to a formatted Excel file. Used when network
           discovery finds new equipment and the user wants it kept in the
           inventory.

    :param printers:
        PT-PT: Impressoras a gravar. / EN-UK: Printers to write.
    :param path:
        PT-PT: Destino. / EN-UK: Destination.
    :return:
        PT-PT: Caminho gravado. / EN-UK: Path written.
    """
    workbook = _new_workbook()
    sheet = workbook[SHEET_NAME]

    for printer in printers:
        sheet.append([
            printer.ip,
            printer.location,
            printer.hostname,
            printer.model,
            printer.serial,
            printer.mac,
            printer.scheme,
            "Sim" if printer.enabled else "Não",
            printer.notes,
        ])

    _style_data_rows(sheet, count=len(printers))

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    _log.info("Inventário gravado: %d impressoras em %s", len(printers), path)
    return path


def create_template(path: Path, overwrite: bool = False) -> Path:
    """
    PT-PT: Cria o modelo Excel que o utilizador vai preencher.

           Leva uma folha de instruções e uma linha de exemplo, porque um
           ficheiro só com cabeçalhos deixa dúvidas sobre o formato esperado —
           sobretudo na coluna Protocolo, que é a que mais confusão gera.

    EN-UK: Creates the Excel template the user will fill in.

           It carries an instructions sheet and one example row, because a file
           with headers alone leaves the expected format in doubt — above all
           in the Protocolo column, which causes the most confusion.

    :param path:
        PT-PT: Destino do modelo. / EN-UK: Destination of the template.
    :param overwrite:
        PT-PT: True substitui um ficheiro existente. Por omissão é False, para
               que a criação automática no arranque nunca apague um inventário
               já preenchido.
        EN-UK: True replaces an existing file. It defaults to False, so that
               automatic creation at start-up can never wipe an inventory that
               has already been filled in.
    :return:
        PT-PT: Caminho do modelo. / EN-UK: Path of the template.
    """
    if path.exists() and not overwrite:
        raise InventoryError(
            f"O ficheiro {path.name} já existe.\n"
            f"Apague-o ou escolha outro nome para criar um modelo novo."
        )

    workbook = _new_workbook()
    _add_instructions_sheet(workbook)
    sheet = workbook[SHEET_NAME]

    # PT-PT: Uma linha de exemplo, marcada como inactiva para não ser
    #        consultada. O utilizador vê o formato esperado e pode apagá-la ou
    #        escrever por cima.
    # EN-UK: One example row, marked inactive so it is never queried. The user
    #        sees the expected format and can delete it or type over it.
    sheet.append([
        "192.168.1.50",
        "Recepção",
        "REC01HP",
        "HP LaserJet MFP E42540",
        "CNBRQ000AA",
        "5C60BA000000",
        "https",
        "Não",
        "Linha de exemplo — apague ou escreva por cima",
    ])

    _style_data_rows(sheet, count=1, example_row=True)

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    _log.info("Modelo de inventário criado em %s", path)
    return path


# ---------------------------------------------------------------------------
# PT-PT: Construção do livro / EN-UK: Workbook construction
# ---------------------------------------------------------------------------


def _new_workbook():
    """
    PT-PT: Cria o livro com a folha de dados formatada e o cabeçalho pronto.
    EN-UK: Creates the workbook with the data sheet formatted and the header
           in place.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.worksheet.datavalidation import DataValidation
    except ImportError as exc:
        raise InventoryError(
            "Falta a biblioteca openpyxl, necessária para criar ficheiros Excel.\n"
            "Execute: pip install -r requirements.txt"
        ) from exc

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = SHEET_NAME

    sheet.append(list(COLUMNS))

    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F5C73")

    for cell in sheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    widths = (16, 24, 20, 30, 20, 18, 12, 10, 40)
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[chr(64 + index)].width = width

    sheet.row_dimensions[1].height = 24

    # PT-PT: Congelar a linha de cabeçalho — numa lista de 24 impressoras
    #        perde-se a noção das colunas ao rolar.
    # EN-UK: Freeze the header row — with 24 printers you lose track of the
    #        columns as soon as you scroll.
    sheet.freeze_panes = "A2"

    # PT-PT: Listas pendentes nas colunas onde só há duas respostas possíveis.
    #        Evita "HTTPS", "Https" e "sim " com espaço, que passariam a
    #        validação mas confundiriam quem lê o ficheiro.
    # EN-UK: Drop-down lists on the columns where only two answers are possible.
    #        This prevents "HTTPS", "Https" and "sim " with a trailing space,
    #        which would pass validation but confuse anyone reading the file.
    protocol_rule = DataValidation(
        type="list", formula1='"http,https"', allow_blank=True,
        prompt="http para a maioria; https nas HP FutureSmart mais recentes.",
        promptTitle="Protocolo",
    )
    active_rule = DataValidation(
        type="list", formula1='"Sim,Não"', allow_blank=True,
        prompt="Não = fica na lista mas não é consultada.",
        promptTitle="Activa",
    )
    sheet.add_data_validation(protocol_rule)
    sheet.add_data_validation(active_rule)
    protocol_rule.add("G2:G500")
    active_rule.add("H2:H500")

    return workbook


def _style_data_rows(sheet, count: int, example_row: bool = False) -> None:
    """
    PT-PT: Aplica tipo de letra e alinhamento às linhas de dados.

    EN-UK: Applies font and alignment to the data rows.

    :param sheet:
        PT-PT: Folha a formatar. / EN-UK: Sheet to format.
    :param count:
        PT-PT: Número de linhas de dados. / EN-UK: Number of data rows.
    :param example_row:
        PT-PT: True apresenta a linha em itálico cinzento, para se perceber à
               primeira vista que é um exemplo e não uma impressora real.
        EN-UK: True renders the row in grey italics, so it is obvious at a
               glance that it is an example and not a real printer.
    """
    from openpyxl.styles import Alignment, Font

    body_font = Font(
        name="Arial", size=10,
        italic=example_row,
        color="808080" if example_row else "000000",
    )

    for row in sheet.iter_rows(min_row=2, max_row=1 + count):
        for cell in row:
            cell.font = body_font
            cell.alignment = Alignment(vertical="center")


def _add_instructions_sheet(workbook) -> None:
    """
    PT-PT: Acrescenta a folha de instruções ao livro.

           Fica em segundo lugar de propósito: quem abre o ficheiro deve ver
           primeiro a lista, e só ir às instruções quando tiver dúvidas.

    EN-UK: Adds the instructions sheet to the workbook.

           It sits second deliberately: whoever opens the file should see the
           list first, and turn to the instructions only when in doubt.

    :param workbook:
        PT-PT: Livro a completar. / EN-UK: Workbook to complete.
    """
    from openpyxl.styles import Alignment, Font

    sheet = workbook.create_sheet(INSTRUCTIONS_SHEET)
    sheet.column_dimensions["A"].width = 22
    sheet.column_dimensions["B"].width = 96

    lines: list[tuple[str, str]] = [
        ("COMO PREENCHER", ""),
        ("", "Preencha uma linha por impressora na folha 'Impressoras'. "
             "Grave o ficheiro e carregue em 'Recarregar' na aplicação."),
        ("", ""),
        ("COLUNAS", ""),
        ("IP", "Obrigatório. Endereço da impressora na rede, por exemplo "
               "10.162.84.144. É o único campo sem o qual a linha é ignorada."),
        ("Localização", "Nome pelo qual as pessoas conhecem a impressora "
                        "(Recepção, Cozinha, Contabilidade). É este nome que "
                        "aparece na aplicação e nos ficheiros PDF gerados."),
        ("Nome de rede", "Opcional. O hostname, se existir."),
        ("Modelo", "Opcional. Preenchido automaticamente pela descoberta na "
                   "rede, quando a impressora o reporta."),
        ("Número de série", "Opcional. Útil para pedidos de garantia."),
        ("MAC", "Opcional. Ajuda a identificar o equipamento se o IP mudar."),
        ("Protocolo", "http ou https. Na dúvida deixe http: a aplicação tenta "
                      "o outro automaticamente se o primeiro falhar. As HP "
                      "FutureSmart mais recentes normalmente só respondem em "
                      "https, com certificado auto-assinado."),
        ("Activa", "Sim ou Não. Ponha Não para uma impressora em reparação: "
                   "continua na lista mas deixa de ser consultada, sem perder "
                   "os dados já preenchidos."),
        ("Notas", "Livre. Por exemplo o contrato de manutenção ou quem é o "
                  "responsável pelo departamento."),
        ("", ""),
        ("DESCOBERTA NA REDE", ""),
        ("", "Em alternativa a preencher isto à mão, use 'Procurar na rede' na "
             "aplicação. Ela varre a gama de endereços que indicar, identifica "
             "as impressoras que encontrar e propõe acrescentá-las a esta "
             "folha, já com o modelo e o número de série preenchidos."),
        ("", ""),
        ("AVISO", ""),
        ("", "Este ficheiro identifica equipamento da rede interna. Trate-o "
             "como documentação técnica interna: não o publique nem o "
             "acrescente a um repositório público."),
    ]

    for label, text in lines:
        sheet.append([label, text])

    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row):
        label_cell, text_cell = row[0], row[1]
        is_heading = bool(label_cell.value) and not text_cell.value

        label_cell.font = Font(
            name="Arial", size=11 if is_heading else 10,
            bold=True, color="1F5C73" if is_heading else "000000",
        )
        label_cell.alignment = Alignment(vertical="top")

        text_cell.font = Font(name="Arial", size=10)
        text_cell.alignment = Alignment(vertical="top", wrap_text=True)
