#!/usr/bin/env python3
"""
PT-PT: Testes do inventário em Excel.

       O inventário é a parte mais crítica desta versão: se a leitura falhar, a
       aplicação fica sem impressoras e não faz nada. Estes testes cobrem os
       ficheiros mal preenchidos, que é o caso normal — um inventário mantido
       por várias pessoas ao longo de meses acumula colunas reordenadas, linhas
       em branco e IP com erros de escrita.

EN-UK: Tests for the Excel inventory.

       The inventory is this version's most critical part: if reading fails, the
       application has no printers and does nothing. These tests cover badly
       filled files, which is the normal case — an inventory maintained by
       several people over months accumulates reordered columns, blank rows and
       mistyped IP addresses.

Created by Redfox using Claude
"""

from __future__ import annotations

from pathlib import Path

import pytest

from tonermon.inventory import (
    COLUMNS,
    InventoryError,
    create_template,
    load,
    save_xlsx,
)
from tonermon.models import Printer


def _write_sheet(path: Path, rows: list[list[object]]) -> Path:
    """
    PT-PT: Escreve uma folha de teste com os cabeçalhos e linhas indicados.
    EN-UK: Writes a test sheet with the given headers and rows.
    """
    from openpyxl import Workbook

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Impressoras"
    for row in rows:
        sheet.append(row)
    workbook.save(path)
    return path


class TestTemplate:
    """
    PT-PT: Criação do modelo Excel.
    EN-UK: Excel template creation.
    """

    def test_creates_file_with_both_sheets(self, tmp_path: Path) -> None:
        """
        PT-PT: O modelo traz a folha de dados e a de instruções.
        EN-UK: The template carries the data sheet and the instructions sheet.
        """
        from openpyxl import load_workbook

        path = create_template(tmp_path / "Impressoras.xlsx")
        workbook = load_workbook(path)

        assert "Impressoras" in workbook.sheetnames
        assert "Instruções" in workbook.sheetnames

    def test_headers_match_expected_columns(self, tmp_path: Path) -> None:
        """
        PT-PT: Os cabeçalhos escritos são exactamente os esperados na leitura.
               Se divergirem, o modelo gerado pela aplicação deixa de poder ser
               lido por ela própria — falha absurda e perfeitamente possível.
        EN-UK: The headers written are exactly those expected on reading. If
               they diverge, the template the application generates can no
               longer be read by the application itself — an absurd failure and
               a perfectly possible one.
        """
        from openpyxl import load_workbook

        path = create_template(tmp_path / "Impressoras.xlsx")
        sheet = load_workbook(path)["Impressoras"]

        assert tuple(cell.value for cell in sheet[1]) == COLUMNS

    def test_example_row_is_inactive(self, tmp_path: Path) -> None:
        """
        PT-PT: A linha de exemplo não pode ser consultada na rede. Um endereço
               fictício activo produziria um erro logo no primeiro arranque.
        EN-UK: The example row must not be queried on the network. An active
               fictitious address would produce an error on the very first run.
        """
        path = create_template(tmp_path / "Impressoras.xlsx")
        printers = load(path)

        assert len(printers) == 1
        assert printers[0].enabled is False

    def test_refuses_to_overwrite_by_default(self, tmp_path: Path) -> None:
        """
        PT-PT: Criar o modelo nunca apaga um inventário existente sem que isso
               seja pedido explicitamente. É a pior falha possível desta
               ferramenta.
        EN-UK: Creating the template never wipes an existing inventory unless
               explicitly asked. It is this tool's worst possible failure.
        """
        path = create_template(tmp_path / "Impressoras.xlsx")

        with pytest.raises(InventoryError):
            create_template(path)

        # PT-PT: Com overwrite explícito, passa a ser permitido.
        # EN-UK: With overwrite set explicitly, it is allowed.
        assert create_template(path, overwrite=True) == path


class TestLoading:
    """
    PT-PT: Leitura de ficheiros de inventário.
    EN-UK: Reading inventory files.
    """

    def test_reads_a_simple_sheet(self, tmp_path: Path) -> None:
        """
        PT-PT: Uma folha bem preenchida é lida na íntegra.
        EN-UK: A properly filled sheet is read in full.
        """
        path = _write_sheet(tmp_path / "p.xlsx", [
            list(COLUMNS),
            ["192.168.1.144", "Purchase", "EXEMPLO01HP", "HP M527",
             "CZCBR1B02F", "7C5758AFBAA3", "https", "Sim", ""],
        ])
        printers = load(path)

        assert len(printers) == 1
        assert printers[0].ip == "192.168.1.144"
        assert printers[0].location == "Purchase"
        assert printers[0].scheme == "https"
        assert printers[0].enabled is True

    def test_columns_may_be_reordered(self, tmp_path: Path) -> None:
        """
        PT-PT: As colunas são procuradas pelo nome, não pela posição. Quem
               mantém a folha vai reordenar as colunas mais cedo ou mais tarde.
        EN-UK: Columns are matched by name, not by position. Whoever maintains
               the sheet will reorder the columns sooner or later.
        """
        path = _write_sheet(tmp_path / "p.xlsx", [
            ["Localização", "Activa", "IP"],
            ["Cozinha", "Sim", "10.0.0.7"],
        ])
        printers = load(path)

        assert printers[0].ip == "10.0.0.7"
        assert printers[0].location == "Cozinha"

    def test_extra_columns_are_ignored(self, tmp_path: Path) -> None:
        """
        PT-PT: Colunas próprias do utilizador (custo, contrato) não estorvam.
        EN-UK: The user's own columns (cost, contract) do not get in the way.
        """
        path = _write_sheet(tmp_path / "p.xlsx", [
            ["IP", "Localização", "Custo anual", "Contrato"],
            ["10.0.0.7", "SPA", "1200", "Xerox"],
        ])
        printers = load(path)

        assert len(printers) == 1
        assert printers[0].location == "SPA"

    def test_alternative_header_names(self, tmp_path: Path) -> None:
        """
        PT-PT: Nomes alternativos são aceites, incluindo em inglês e sem
               acentos, porque uma lista exportada de outra ferramenta raramente
               usa os nomes exactos.
        EN-UK: Alternative names are accepted, including English and unaccented
               ones, because a list exported from another tool rarely uses the
               exact names.
        """
        path = _write_sheet(tmp_path / "p.xlsx", [
            ["Address", "Location", "Model", "S/N"],
            ["10.0.0.7", "IT", "HP E50145", "PHCBQBS0K1"],
        ])
        printers = load(path)

        assert printers[0].location == "IT"
        assert printers[0].model == "HP E50145"
        assert printers[0].serial == "PHCBQBS0K1"

    def test_invalid_ip_is_skipped(self, tmp_path: Path) -> None:
        """
        PT-PT: Uma linha com IP errado é ignorada em vez de rebentar. Validar
               aqui evita que o erro de escrita apareça mais tarde como um
               timeout de rede sem explicação.
        EN-UK: A row with a bad IP is skipped rather than crashing. Validating
               here stops the typo surfacing later as an unexplained network
               timeout.
        """
        path = _write_sheet(tmp_path / "p.xlsx", [
            ["IP", "Localização"],
            ["10.0.0.7", "Boa"],
            ["10.0.0.999", "IP impossível"],
            ["nao-e-um-ip", "Texto"],
        ])
        printers = load(path)

        assert [printer.location for printer in printers] == ["Boa"]

    def test_blank_rows_are_skipped(self, tmp_path: Path) -> None:
        """
        PT-PT: Linhas em branco no fim da folha são normais e não geram avisos.
        EN-UK: Blank rows at the end of the sheet are normal and raise no
               warnings.
        """
        path = _write_sheet(tmp_path / "p.xlsx", [
            ["IP", "Localização"],
            ["10.0.0.7", "Boa"],
            [None, None],
            ["", ""],
        ])
        assert len(load(path)) == 1

    def test_activa_accepts_several_spellings(self, tmp_path: Path) -> None:
        """
        PT-PT: A coluna Activa aceita várias formas de "sim". Uma lista mantida
               por várias pessoas terá "Sim", "SIM", "x" e "yes" ao mesmo tempo.
        EN-UK: The Activa column accepts several forms of "yes". A list
               maintained by several people will hold "Sim", "SIM", "x" and
               "yes" all at once.
        """
        path = _write_sheet(tmp_path / "p.xlsx", [
            ["IP", "Activa"],
            ["10.0.0.1", "Sim"],
            ["10.0.0.2", "SIM"],
            ["10.0.0.3", "yes"],
            ["10.0.0.4", "x"],
            ["10.0.0.5", "Não"],
            ["10.0.0.6", "no"],
        ])
        printers = load(path)

        assert [printer.enabled for printer in printers] == [
            True, True, True, True, False, False
        ]

    def test_unknown_protocol_falls_back_to_http(self, tmp_path: Path) -> None:
        """
        PT-PT: Um protocolo inválido não impede a leitura — a aplicação tenta
               os dois de qualquer forma.
        EN-UK: An invalid protocol does not block reading — the application
               tries both anyway.
        """
        path = _write_sheet(tmp_path / "p.xlsx", [
            ["IP", "Protocolo"],
            ["10.0.0.7", "ftp"],
        ])
        assert load(path)[0].scheme == "http"

    def test_missing_ip_column_is_an_error(self, tmp_path: Path) -> None:
        """
        PT-PT: Sem coluna IP o ficheiro é inútil, e a mensagem tem de dizer
               quais são as colunas esperadas.
        EN-UK: Without an IP column the file is useless, and the message must
               say which columns are expected.
        """
        path = _write_sheet(tmp_path / "p.xlsx", [
            ["Localização", "Modelo"],
            ["Cozinha", "HP"],
        ])
        with pytest.raises(InventoryError, match="IP"):
            load(path)

    def test_missing_file_is_an_error(self, tmp_path: Path) -> None:
        """
        PT-PT: Um ficheiro inexistente dá uma mensagem útil, não um traceback.
        EN-UK: A missing file gives a useful message, not a traceback.
        """
        with pytest.raises(InventoryError):
            load(tmp_path / "nao-existe.xlsx")

    def test_unsupported_extension_is_an_error(self, tmp_path: Path) -> None:
        """
        PT-PT: Formatos não suportados são recusados com clareza.
        EN-UK: Unsupported formats are refused clearly.
        """
        path = tmp_path / "lista.docx"
        path.write_text("nada", encoding="utf-8")

        with pytest.raises(InventoryError):
            load(path)


class TestCsv:
    """
    PT-PT: Leitura de CSV, para quem já tenha a lista nesse formato.
    EN-UK: CSV reading, for anyone whose list is already in that format.
    """

    def test_reads_comma_separated(self, tmp_path: Path) -> None:
        """
        PT-PT: CSV com vírgulas, o formato de uma lista feita à mão.
        EN-UK: Comma-separated CSV, the format of a hand-made list.
        """
        path = tmp_path / "p.csv"
        path.write_text(
            "IP,Localização,Activa\n10.0.0.7,Cozinha,Sim\n", encoding="utf-8"
        )
        printers = load(path)

        assert printers[0].location == "Cozinha"

    def test_reads_semicolon_separated(self, tmp_path: Path) -> None:
        """
        PT-PT: CSV com ponto e vírgula, que é o que o Excel português grava.
        EN-UK: Semicolon-separated CSV, which is what Portuguese Excel saves.
        """
        path = tmp_path / "p.csv"
        path.write_text(
            "IP;Localização;Activa\n10.0.0.7;Recepção;Sim\n", encoding="utf-8"
        )
        printers = load(path)

        assert printers[0].location == "Recepção"

    def test_strips_excel_byte_order_mark(self, tmp_path: Path) -> None:
        """
        PT-PT: O Excel grava CSV com marca de ordem de bytes, que sem
               tratamento contaminaria o primeiro cabeçalho e faria a coluna IP
               deixar de ser reconhecida.
        EN-UK: Excel saves CSV with a byte-order mark which, left untreated,
               would contaminate the first header and stop the IP column being
               recognised.
        """
        path = tmp_path / "p.csv"
        path.write_text(
            "IP;Localização\n10.0.0.7;Bar\n", encoding="utf-8-sig"
        )
        assert load(path)[0].ip == "10.0.0.7"


class TestSaving:
    """
    PT-PT: Escrita do inventário.
    EN-UK: Writing the inventory.
    """

    def test_round_trips(self, tmp_path: Path) -> None:
        """
        PT-PT: O que é gravado é lido de volta igual, acentos incluídos. Este é
               o teste que protege o caminho da descoberta na rede, que grava o
               ficheiro por cima do que o utilizador mantém.
        EN-UK: What is written reads back identically, accents included. This is
               the test protecting the network-discovery path, which writes the
               file over what the user maintains.
        """
        original = [
            Printer(ip="10.0.0.7", location="Recepção Ala Sul",
                    model="HP E50145", scheme="https", enabled=True),
            Printer(ip="10.0.0.8", location="Direcção",
                    notes="Em reparação", enabled=False),
        ]
        path = save_xlsx(original, tmp_path / "out.xlsx")
        restored = load(path)

        assert len(restored) == 2
        assert restored[0].location == "Recepção Ala Sul"
        assert restored[0].scheme == "https"
        assert restored[1].enabled is False
        assert restored[1].notes == "Em reparação"

    def test_creates_missing_folders(self, tmp_path: Path) -> None:
        """
        PT-PT: Gravar numa pasta que ainda não existe funciona.
        EN-UK: Saving into a folder that does not yet exist works.
        """
        destination = tmp_path / "a" / "b" / "out.xlsx"
        save_xlsx([Printer(ip="10.0.0.7")], destination)

        assert destination.is_file()
