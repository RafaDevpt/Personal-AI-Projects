#!/usr/bin/env python3
"""
PT-PT: Inventário de equipamentos.

       A lista de switches de uma casa já existe algures — quase sempre num
       Excel que alguém mantém à mão. Este módulo lê esse Excel e escreve-o de
       volta, em vez de obrigar a manter a mesma lista duas vezes.

       O que não está aqui, e nunca vai estar, são credenciais. Um inventário é
       um ficheiro que anda por email, que fica numa pasta partilhada e que
       acaba anexado a um ticket. Guardar palavras-passe nele seria criar o
       problema que a ferramenta devia ajudar a evitar. As credenciais são
       pedidas quando a sessão começa e desaparecem quando ela acaba.

EN-UK: Device inventory.

       A property's switch list already exists somewhere — nearly always in a
       spreadsheet somebody maintains by hand. This module reads that
       spreadsheet and writes it back, rather than forcing the same list to be
       kept twice.

       What is not here, and never will be, are credentials. An inventory is a
       file that travels by e-mail, sits in a shared folder and ends up
       attached to a ticket. Storing passwords in it would create the very
       problem the tool should help avoid. Credentials are asked for when the
       session starts and vanish when it ends.

Created by Redfox using Claude
"""

from __future__ import annotations

import csv
import json
import logging
from pathlib import Path
from typing import Any

from .models import Device, Platform

logger = logging.getLogger(__name__)

# PT-PT: Cabeçalhos da folha de inventário, pela ordem em que são escritos.
# EN-UK: Inventory sheet headers, in the order they are written.
COLUMNS = ["Nome", "Endereco", "Plataforma", "Modelo", "Local", "Porta", "Notas"]

# PT-PT: Nomes de plataforma aceites na importação, incluindo o que uma pessoa
#        escreveria à mão numa folha de cálculo.
# EN-UK: Platform names accepted on import, including what a person would type
#        by hand into a spreadsheet.
_PLATFORM_ALIASES: dict[str, Platform] = {
    "aruba": Platform.ARUBA_CX,
    "aruba_cx": Platform.ARUBA_CX,
    "aruba cx": Platform.ARUBA_CX,
    "aos-cx": Platform.ARUBA_CX,
    "aos cx": Platform.ARUBA_CX,
    "cisco": Platform.CISCO_IOS,
    "cisco_ios": Platform.CISCO_IOS,
    "ios": Platform.CISCO_IOS,
    "ios-xe": Platform.CISCO_IOS,
    "catalyst": Platform.CISCO_IOS,
    "edgeswitch": Platform.UBIQUITI_EDGESWITCH,
    "ubiquiti_edgeswitch": Platform.UBIQUITI_EDGESWITCH,
    "edge": Platform.UBIQUITI_EDGESWITCH,
    "unifi": Platform.UBIQUITI_UNIFI,
    "ubiquiti_unifi": Platform.UBIQUITI_UNIFI,
    "usw": Platform.UBIQUITI_UNIFI,
}


class InventoryError(ValueError):
    """
    PT-PT: Inventário ilegível, com a razão em português.
    EN-UK: Unreadable inventory, with the reason in Portuguese.
    """


def parse_platform(value: str, default: Platform = Platform.ARUBA_CX) -> Platform:
    """
    PT-PT: Interpreta o que estiver escrito na coluna da plataforma.

    EN-UK: Interprets whatever is written in the platform column.

    :param value:
        PT-PT: Texto da célula. / EN-UK: Cell text.
    :param default:
        PT-PT: O que devolver se a célula estiver vazia.
        EN-UK: What to return when the cell is empty.
    :return:
        PT-PT: Plataforma correspondente. / EN-UK: Matching platform.
    :raises InventoryError:
        PT-PT: Se o texto não corresponder a nenhuma plataforma conhecida.
        EN-UK: If the text matches no known platform.
    """
    texto = (value or "").strip().lower()
    if not texto:
        return default
    if texto in _PLATFORM_ALIASES:
        return _PLATFORM_ALIASES[texto]
    try:
        return Platform(texto)
    except ValueError as exc:
        conhecidas = ", ".join(sorted({p.value for p in Platform}))
        raise InventoryError(f"Plataforma desconhecida: {value!r}. Conhecidas: {conhecidas}") from exc


def to_rows(devices: list[Device]) -> list[list[Any]]:
    """
    PT-PT: Converte os equipamentos em linhas, com cabeçalho.
    EN-UK: Turns the devices into rows, header included.
    """
    linhas: list[list[Any]] = [list(COLUMNS)]
    for device in devices:
        linhas.append(
            [
                device.name,
                device.host,
                device.platform.value,
                device.model,
                device.site,
                device.port,
                device.notes,
            ]
        )
    return linhas


def from_rows(rows: list[list[Any]]) -> list[Device]:
    """
    PT-PT: Constrói os equipamentos a partir das linhas de uma folha.

           A primeira linha é tratada como cabeçalho apenas se a primeira
           célula for "Nome" — uma folha exportada e uma folha escrita à mão
           chegam ambas aqui, e nem sempre com cabeçalho.

    EN-UK: Builds the devices from a sheet's rows.

           The first row is treated as a header only when the first cell reads
           "Nome" — both an exported sheet and a hand-written one land here,
           and not always with a header.

    :param rows:
        PT-PT: Linhas da folha. / EN-UK: Sheet rows.
    :return:
        PT-PT: Equipamentos lidos, pela ordem da folha.
        EN-UK: Devices read, in sheet order.
    :raises InventoryError:
        PT-PT: Se uma linha não tiver nome ou endereço.
        EN-UK: If a row has no name or no address.
    """
    devices: list[Device] = []
    for index, row in enumerate(rows, start=1):
        celulas = [("" if c is None else str(c)).strip() for c in row]
        if not any(celulas):
            continue
        if index == 1 and celulas and celulas[0].lower() == COLUMNS[0].lower():
            continue

        celulas += [""] * (len(COLUMNS) - len(celulas))
        nome, endereco, plataforma, modelo, local, porta, notas = celulas[: len(COLUMNS)]

        if not nome or not endereco:
            raise InventoryError(f"Linha {index}: faltam o nome ou o endereço.")

        devices.append(
            Device(
                name=nome,
                host=endereco,
                platform=parse_platform(plataforma),
                model=modelo,
                site=local,
                port=int(porta) if porta.isdigit() else 22,
                notes=notas,
            )
        )
    return devices


# ---------------------------------------------------------------------------
# PT-PT: JSON — o formato que a aplicação usa por omissão.
# EN-UK: JSON — the format the application uses by default.
# ---------------------------------------------------------------------------


def save_json(devices: list[Device], path: Path) -> Path:
    """PT-PT: Grava o inventário em JSON. / EN-UK: Writes the inventory as JSON."""
    path.parent.mkdir(parents=True, exist_ok=True)
    dados = [
        {
            "name": d.name,
            "host": d.host,
            "platform": d.platform.value,
            "model": d.model,
            "site": d.site,
            "port": d.port,
            "notes": d.notes,
        }
        for d in devices
    ]
    path.write_text(json.dumps(dados, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
    return path


def load_json(path: Path) -> list[Device]:
    """
    PT-PT: Lê o inventário em JSON. Um ficheiro que ainda não exista devolve
           uma lista vazia — é o estado normal na primeira execução.
    EN-UK: Reads the JSON inventory. A file that does not exist yet returns an
           empty list — that is the normal state on first run.
    """
    if not path.exists():
        return []
    try:
        dados = json.loads(path.read_text(encoding="utf-8"))
    except json.JSONDecodeError as exc:
        raise InventoryError(f"O inventário {path.name} não é JSON válido: {exc}") from exc
    if not isinstance(dados, list):
        raise InventoryError(f"O inventário {path.name} devia conter uma lista.")

    devices: list[Device] = []
    for entrada in dados:
        devices.append(
            Device(
                name=str(entrada.get("name") or ""),
                host=str(entrada.get("host") or ""),
                platform=parse_platform(str(entrada.get("platform") or "")),
                model=str(entrada.get("model") or ""),
                site=str(entrada.get("site") or ""),
                port=int(entrada.get("port") or 22),
                notes=str(entrada.get("notes") or ""),
            )
        )
    return devices


# ---------------------------------------------------------------------------
# PT-PT: CSV e Excel — para as listas que já existem.
# EN-UK: CSV and Excel — for the lists that already exist.
# ---------------------------------------------------------------------------


def load_csv(path: Path) -> list[Device]:
    """PT-PT: Lê um inventário em CSV. / EN-UK: Reads a CSV inventory."""
    with path.open(encoding="utf-8-sig", newline="") as handle:
        # PT-PT: O Excel português grava CSV com ponto e vírgula.
        # EN-UK: Portuguese Excel writes CSV with semicolons.
        amostra = handle.read(4096)
        handle.seek(0)
        delimitador = ";" if amostra.count(";") > amostra.count(",") else ","
        return from_rows([list(row) for row in csv.reader(handle, delimiter=delimitador)])


def load_xlsx(path: Path) -> list[Device]:
    """
    PT-PT: Lê um inventário em Excel, da primeira folha.

    EN-UK: Reads an Excel inventory, from the first sheet.

    :raises InventoryError:
        PT-PT: Se o openpyxl não estiver instalado.
        EN-UK: If openpyxl is not installed.
    """
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise InventoryError(
            "O openpyxl não está instalado. Instale com: pip install openpyxl"
        ) from exc

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook.worksheets[0]
        return from_rows([list(row) for row in sheet.iter_rows(values_only=True)])
    finally:
        workbook.close()


def save_xlsx(devices: list[Device], path: Path) -> Path:
    """
    PT-PT: Escreve o inventário em Excel, com o cabeçalho a negrito e as
           colunas com largura utilizável — uma folha que abre com tudo
           encostado não é lida por ninguém.

    EN-UK: Writes the inventory to Excel, with a bold header and usable column
           widths — a sheet that opens with everything squashed gets read by
           nobody.

    :raises InventoryError:
        PT-PT: Se o openpyxl não estiver instalado.
        EN-UK: If openpyxl is not installed.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
    except ImportError as exc:
        raise InventoryError(
            "O openpyxl não está instalado. Instale com: pip install openpyxl"
        ) from exc

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Equipamentos"

    for linha in to_rows(devices):
        sheet.append(linha)
    for cell in sheet[1]:
        cell.font = Font(bold=True)
    for indice, largura in enumerate([22, 16, 22, 20, 18, 8, 40], start=1):
        sheet.column_dimensions[sheet.cell(row=1, column=indice).column_letter].width = largura

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    return path


def create_template(path: Path) -> Path:
    """
    PT-PT: Grava uma folha de exemplo, com uma linha por plataforma suportada
           para servir de referência de escrita.

    EN-UK: Writes a sample sheet, with one row per supported platform to serve
           as a writing reference.

    :param path:
        PT-PT: Destino do modelo. / EN-UK: Template destination.
    :return:
        PT-PT: O caminho gravado. / EN-UK: The written path.
    """
    exemplos = [
        Device("SW-PISO1-01", "10.0.10.11", Platform.ARUBA_CX, "6300M", "Piso 1", notes="Bastidor A"),
        Device("SW-CORE-01", "10.0.10.1", Platform.CISCO_IOS, "C9300", "Datacenter"),
        Device("SW-COZINHA", "10.0.10.31", Platform.UBIQUITI_EDGESWITCH, "ES-24-250W", "Cozinha"),
        Device("SW-LOBBY", "10.0.10.41", Platform.UBIQUITI_UNIFI, "USW-24-PoE", "Lobby",
               notes="Gerido pelo controlador"),
    ]
    return save_xlsx(exemplos, path)


def load(path: Path) -> list[Device]:
    """
    PT-PT: Lê o inventário escolhendo o leitor pela extensão.

    EN-UK: Reads the inventory, picking the reader by extension.

    :param path:
        PT-PT: Ficheiro .json, .csv, .xlsx ou .xlsm.
        EN-UK: A .json, .csv, .xlsx or .xlsm file.
    :return:
        PT-PT: Equipamentos. / EN-UK: Devices.
    :raises InventoryError:
        PT-PT: Se a extensão não for reconhecida.
        EN-UK: If the extension is not recognised.
    """
    sufixo = path.suffix.lower()
    if sufixo == ".json":
        return load_json(path)
    if sufixo == ".csv":
        return load_csv(path)
    if sufixo in {".xlsx", ".xlsm"}:
        return load_xlsx(path)
    raise InventoryError(f"Não sei ler ficheiros {sufixo or 'sem extensão'}.")
